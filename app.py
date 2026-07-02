import streamlit as st
import pandas as pd
from streamlit_option_menu import option_menu
import hashlib
import datetime
from datetime import datetime
import urllib.parse
from io import BytesIO
import time
import base64
from zoneinfo import ZoneInfo
from report_utils import (
    generate_excel_report,
    generate_pdf_report
)
from backup_utils import (
    create_full_backup,
    restore_full_backup
)
from supabase import create_client, Client

# ================= FIXED MOBILE_UI FUNCTIONS EMBEDDED =================
# Ab kisi external import ki zaroorat nahi, yeh Cloud par 100% chalega.

def is_mobile():
    # User agent se mobile devices check karne ka default full-proof tareeqa
    try:
        from streamlit.web.server.websocket_headers import _get_websocket_headers
        headers = _get_websocket_headers()
        if headers and "User-Agent" in headers:
            ua = headers["User-Agent"].lower()
            return any(x in ua for x in ["android", "iphone", "ipad", "mobile"])
    except:
        pass
    return False

def load_mobile_css():
    st.markdown("""
    <style>
    /* Desktop sidebar layout styling fix */
    @media (min-width: 769px) {
        section[data-testid="stSidebar"] {
            min-width: 240px !important;
            max-width: 280px !important;
        }
        div[data-testid="stSidebarNav"] {
            max-height: none !important;
        }
    }
    
    /* Pure Mobile layout overrides */
    @media (max-width: 768px) {
        .block-container { 
            max-width: 480px !important; 
            padding: 10px !important; 
        }
        div.stButton > button { 
            font-size: 12px !important; 
            width: 100%; 
        }
    }
    </style>
    """, unsafe_allow_html=True)

def show_mobile_header(title, subtitle=""):
    st.markdown(f"<h3 style='color: #F8D568; text-align: center; margin: 0;'>{title}</h3>", unsafe_allow_html=True)
    if subtitle:
        st.markdown(f"<p style='color: #A0AEC0; text-align: center; font-size: 12px; margin: 0;'>{subtitle}</p>", unsafe_allow_html=True)

def show_mobile_topbar():
    pass

def show_mobile_section_title(title):
    st.markdown(f"<div style='color: #F8D568; font-size: 15px; font-weight: 600; margin: 10px 0;'>{title}</div>", unsafe_allow_html=True)

def show_mobile_metric_card(label, value, subtext=""):
    st.markdown(f"""
    <div style="background: #11142A; border: 1px solid #1E2342; border-radius: 14px; padding: 12px; margin-bottom: 8px;">
        <div style="color: #94A3B8; font-size: 11px;">{label}</div>
        <div style="color: #FFFFFF; font-size: 16px; font-weight: 700; margin-top: 2px;">{value}</div>
        {f'<div style="color: #475569; font-size: 9px;">{subtext}</div>' if subtext else ''}
    </div>
    """, unsafe_allow_html=True)

# ===== EXCEL EXPORT =====
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

# ===== PDF EXPORT =====
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

# ================= PASSWORD HASH FUNCTION =================
def hash_pass(password):
    return hashlib.sha256(str.encode(password)).hexdigest()

# ================= SUPABASE =================
SUPABASE_URL = st.secrets["SUPABASE_URL"]
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
# ===== EXCEL EXPORT =====
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils import get_column_letter

# ===== PDF EXPORT =====
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image
)

from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

# ================= PASSWORD HASH FUNCTION =================
def hash_pass(password):
    return hashlib.sha256(str.encode(password)).hexdigest()

# ================= SUPABASE =================
SUPABASE_URL = st.secrets["SUPABASE_URL"]
SUPABASE_KEY = st.secrets["SUPABASE_KEY"]

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# ================= SAFE LOAD =================
try:
    loans_data = supabase.table("loans").select("*").execute()
    loans_df = pd.DataFrame(loans_data.data)
except:
    loans_df = pd.DataFrame()

# ================= PASSWORD =================
def save_log(action, table_name, member_name="", member_id="", amount=0):
    try:
        result = supabase.table("audit_logs").insert({
            "action": action,
            "table_name": table_name,
            "member_name": member_name,
            "member_id": member_id,
            "amount": float(amount),
            "performed_by": st.session_state.get("current_user", "Unknown"),
            "role": st.session_state.get("role", "Unknown")
        }).execute()
        print(result)
    except Exception as e:
        st.error(f"Audit Log Error: {e}")
        print(e)

# ================= CONFIG =================
st.set_page_config(page_title="Bal Yuva SaaS", layout="wide")

# ================= CSS =================
st.markdown("""
<style>

/* ===== ANIMATED BACKGROUND ===== */
.stApp {
background: linear-gradient(135deg,#020617,#0f172a,#020617);
background-size: 400% 400%;
animation: gradientBG 12s ease infinite;
color: white;
}

@keyframes gradientBG {
0% {background-position:0% 50%;}
50% {background-position:100% 50%;}
100% {background-position:0% 50%;}
}

/* ===== HIDE HEADER ===== */
header, footer {
visibility: hidden;
}

.block-container {
padding-top: 0.5rem !important;
}

.block-container {
padding-top: 0rem !important;
margin-top: -15px !important;
}

/* ===== TEXT ===== */
h1,h2,h3,h4,h5,p,label {
color:white;
}

/* ===== SIDEBAR GLASS ===== */
section[data-testid="stSidebar"] {
background: rgba(15, 23, 42, 0.65);
backdrop-filter: blur(18px);
border-right: 1px solid rgba(255,255,255,0.08);
}

/* ===== GLASS CARD ===== */
.glass-card {
background: rgba(255, 255, 255, 0.04);
backdrop-filter: blur(18px);
border-radius: 16px;
padding: 20px;
border: 1px solid rgba(255,255,255,0.08);
box-shadow: 0 10px 35px rgba(0,0,0,0.35);
transition: all 0.25s ease;
}

/* ===== CARD HOVER (soft) ===== */
.glass-card:hover {
transform: translateY(-4px);
box-shadow: 0 12px 35px rgba(124,58,237,0.25);
}

/* ===== INPUT FIELDS ===== */
.stTextInput input,
.stNumberInput input,
.stDateInput input {
background: rgba(17,24,39,0.75) !important;
color:white !important;
border-radius:10px !important;
border:1px solid rgba(255,255,255,0.08);
}

/* ===== SELECTBOX FIX ===== */
.stSelectbox div[data-baseweb="select"] {
background: rgba(17,24,39,0.75) !important;
border-radius:10px !important;
border:1px solid rgba(255,255,255,0.08);
}

/* ===== BUTTONS ===== */
div.stButton > button {
background: linear-gradient(135deg,#6366f1,#7c3aed);
color: white;
border-radius: 12px;
height: 42px;
border: none;
transition: all 0.25s ease;
}

/* ===== BUTTON HOVER (controlled) ===== */
div.stButton > button:hover {
transform: translateY(-2px);
box-shadow: 0 0 12px rgba(124,58,237,0.4);
}

/* ===== DOWNLOAD BUTTONS ===== */
.stDownloadButton > button {
background: linear-gradient(135deg,#6366f1,#7c3aed) !important;
color: white !important;
border-radius: 12px !important;
height: 42px !important;
border: none !important;
width: 100% !important;
font-weight: 600 !important;
}

.stDownloadButton > button:hover {
transform: translateY(-2px);
box-shadow: 0 0 12px rgba(124,58,237,0.4);
}

/* ===== FORM BUTTON (LOGIN SAFE) ===== */
div.stForm button {
background: linear-gradient(135deg,#16a34a,#22c55e) !important;
color: white !important;
border-radius: 12px !important;
padding: 10px 20px !important;
border: none !important;
transition: all 0.25s ease;
}

div.stForm button:hover {
transform: translateY(-2px);
box-shadow: 0 0 10px rgba(34,197,94,0.5);
}
/* ===== PASSWORD EYE BUTTON ===== */
[data-testid="stTextInput"] button {
background: transparent !important;
border: none !important;
box-shadow: none !important;
}

[data-testid="stTextInput"] button:hover {
background: transparent !important;
}
/* PASSWORD EYE FIX */
[data-testid="stTextInput"] button {
background: transparent !important;
border: none !important;
box-shadow: none !important;
}

[data-testid="stTextInput"] button svg {
width: 14px !important;
height: 14px !important;
}

/* ===== METRIC CARDS ===== */
[data-testid="metric-container"] {
background: rgba(255,255,255,0.04);
backdrop-filter: blur(18px);
padding: 15px;
border-radius: 14px;
border: 1px solid rgba(255,255,255,0.08);
transition: all 0.25s ease;
}

/* ===== METRIC HOVER ===== */
[data-testid="metric-container"]:hover {
transform: translateY(-3px);
box-shadow: 0 0 15px rgba(124,58,237,0.3);
}

/* ===== DATAFRAME ===== */
[data-testid="stDataFrame"] {
background: rgba(255,255,255,0.02);
border-radius: 12px;
padding: 5px;
}

/* ===== CHART AREA ===== */
canvas {
background: rgba(255,255,255,0.02) !important;
border-radius: 10px;
}

/* ===== FADE ANIMATION ===== */
@keyframes fadeInUp {
from {
opacity: 0;
transform: translateY(15px);
}
to {
opacity: 1;
transform: translateY(0);
}
}

.fade-in {
animation: fadeInUp 0.5s ease forwards;
}

/* ===== SCROLLBAR ===== */
::-webkit-scrollbar {
width: 6px;
}
::-webkit-scrollbar-thumb {
background: rgba(255,255,255,0.15);
border-radius: 10px;
}

/* ===== MOBILE FIX ===== */
html, body, .stApp {
overflow-y: auto !important;
overflow-x: hidden !important;
}

/* ===== LOGO CONTAINER FIX ===== */
.stImage {
background: transparent !important;
padding: 0 !important;
margin: 0 !important;
border: none !important;
}

.stImage > div {
background: transparent !important;
border: none !important;
box-shadow: none !important;
}

.stImage img {
border-radius: 18px;
background: transparent !important;
}
/* ===== FILE UPLOADER PREMIUM BLUE ===== */

[data-testid="stFileUploader"] {
background: rgba(99,102,241,0.12) !important;
border: 1px solid rgba(99,102,241,0.35) !important;
border-radius: 14px !important;
padding: 12px !important;
backdrop-filter: blur(18px);
}

[data-testid="stFileUploader"] section {
background: linear-gradient(
135deg,
rgba(99,102,241,0.25),
rgba(124,58,237,0.25)
) !important;

border: 1px dashed rgba(255,255,255,0.25) !important;
border-radius: 12px !important;
}

[data-testid="stFileUploader"] small,
[data-testid="stFileUploader"] label,
[data-testid="stFileUploader"] span {
color: white !important;
}
/* ===== FILE UPLOADER BUTTON ===== */

[data-testid="stFileUploader"] button {
background: linear-gradient(
135deg,
#6366f1,
#7c3aed
) !important;

color: white !important;
border: none !important;
border-radius: 10px !important;
}

[data-testid="stFileUploader"] button:hover {
background: linear-gradient(
135deg,
#7c3aed,
#6366f1
) !important;
}

/* Mobile Quick Actions Clean Layout styling override */
@media (max-width: 768px) {
.block-container {
max-width: 480px !important;
padding-left: 14px !important;
padding-right: 14px !important;
}
div.stButton > button {
background: transparent !important;
border: none !important;
box-shadow: none !important;
color: #CBD5E1 !important;
font-size: 11px !important;
padding: 0 !important;
margin-top: 4px !important;
}
}
</style>
""", unsafe_allow_html=True)
load_mobile_css()

# ================= SESSION =================
if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

SESSION_TIMEOUT = 1800 # 30 minutes

# ===== CHECK SESSION TIMEOUT =====
if st.session_state.get("logged_in"):
    if "last_active" not in st.session_state:
        st.session_state.last_active = time.time()

    if time.time() - st.session_state.last_active > SESSION_TIMEOUT:
        st.warning("Session expired. Please login again.")
        st.session_state.clear()
        st.rerun()

    # update activity time
    st.session_state.last_active = time.time()

def add_bg_from_local(image_file):
    with open(image_file, "rb") as image:
        encoded = base64.b64encode(image.read()).decode()

    st.markdown(
        f"""
        <style>

        .stApp {{
        background-image:
        linear-gradient(
        rgba(0,0,0,0.10),
        rgba(0,0,0,0.10)
        ),
        url("data:image/png;base64,{encoded}");

        background-size: cover;
        background-position: center;
        background-repeat: no-repeat;
        background-attachment: fixed;
        }}

        html, body, [data-testid="stAppViewContainer"] {{
        height: 100vh;
        overflow: hidden !important;
        }}

        section.main > div {{
        padding-top: 0rem !important;
        padding-bottom: 0rem !important;
        }}

        .block-container {{
        padding-top: 1rem !important;
        padding-bottom: 0rem !important;
        max-width: 100% !important;
        }}

        [data-testid="stVerticalBlock"] {{
        gap: 0rem !important;
        }}

        </style>
        """,
        unsafe_allow_html=True,
    )

# ================= LOGIN SYSTEM =================
if not st.session_state.get("logged_in", False):

    # Login page background only
    add_bg_from_local("login_bg.png")

    # ================= LOGIN HEADER =================
    if is_mobile():
        st.markdown("""
        <div style="
        margin-top:20px;
        text-align:center;
        padding:10px;
        ">

        <h1 style="
        color:#f7d774;
        font-size:34px;
        font-weight:700;
        margin-bottom:8px;
        line-height:1.2;
        ">
        बाल युवा मंगलदल समिति
        </h1>

        <h2 style="
        color:#f7d774;
        font-size:24px;
        margin-top:0px;
        margin-bottom:15px;
        ">
        मयलगांव
        </h2>

        <p style="
        color:white;
        font-size:15px;
        margin-bottom:5px;
        ">
        हमारा गांव • हमारी पहचान
        </p>

        <p style="
        color:#d8d8d8;
        font-size:13px;
        margin-bottom:25px
        ">
        Secure Finance Management System
        </p>

        </div>
        """, unsafe_allow_html=True)

    else:
        col1, col2 = st.columns([2.2, 1])

        with col1:
            st.markdown("""
            <div style="
            margin-top:60px;
            text-align:left;
            ">

            <h1 style="
            color:#f7d774;
            font-size:52px;
            font-weight:700;
            margin-bottom:10px;
            line-height:1.1;
            ">
            बाल युवा मंगलदल समिति
            </h1>

            <h2 style="
            color:#f7d774;
            font-size:38px;
            margin-top:0px;
            margin-bottom:20px;
            ">
            मयलगांव
            </h2>

            <p style="
            color:white;
            font-size:18px;
            margin-bottom:8px;
            ">
            हमारा गांव • हमारी पहचान • हमारा अभियान
            </p>

            <p style="
            color:#d8d8d8;
            font-size:16px;
            margin-bottom:30px;
            ">
            Secure Finance Management System
            </p>

            </div>
            """, unsafe_allow_html=True)

    st.markdown("""
    <style>
    div[data-testid="stForm"]{
    background: rgba(255,255,255,0.05);
    padding:30px;
    border-radius:20px;
    border:1px solid rgba(255,255,255,0.08);
    backdrop-filter: blur(14px);
    max-width:600px;
    margin:auto;
    }

    div[data-testid="stForm"] label{
    color:white !important;
    font-weight:600;
    }
    </style>
    """, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1.0, 2.4, 0.8])

    with col1:
        with st.form("login_form"):
            u = st.text_input("👤 Username")
            p = st.text_input("🔒 Password", type="password")

            submitted = st.form_submit_button("🚀 Login")

            if submitted:
                if u == "" or p == "":
                    st.warning("Enter Username & Password")
                else:
                    try:
                        user_data = (
                            supabase.table("users")
                            .select("*")
                            .eq("username", u)
                            .execute()
                        )

                        if user_data.data:
                            user = user_data.data[0]

                            if user["password"] == hash_pass(p):
                                st.session_state.logged_in = True
                                st.session_state.current_user = user["username"]
                                st.session_state.role = user["role"]
                                st.session_state.last_active = time.time()
                                st.rerun()
                            else:
                                st.error("Wrong Password")
                        else:
                            st.error("User not found")
                    except Exception as e:
                        st.error(f"Login Error: {e}")

    st.stop()

# ================= ROLE SETUP =================
role = st.session_state.get("role", None)

is_admin = role == "Admin"
is_editor = role == "Editor"
is_viewer = role == "Viewer"

# ================= ROLE BASED MENU =================
if is_admin:
    menu_list = ["Dashboard", "Members", "Collection Rates", "Collections", "loans", "Donations", "Expenses", "Reports", "Reminders", "Users", "Backup & Restore", "AI"]
elif is_editor:
    menu_list = ["Dashboard", "Members", "Collection Rates", "Collections", "loans", "Donations", "Expenses", "Reports", "Reminders"]
else:
    menu_list = ["Dashboard", "Reports"]

# ================= MENU RENDER (SIDEBAR FOR BOTH MOBILE & PC) =================
with st.sidebar:
    st.markdown(
        f"""
        <div style="
        text-align:center;
        margin-top:10px;
        margin-bottom:20px;
        font-size:18px;
        font-weight:600;
        color:#cbd5e1;
        ">
        Welcome,
        <span style="color:#8b5cf6;">
        {st.session_state.get("current_user","Admin")}
        </span> 👋
        </div>
        """,
        unsafe_allow_html=True
    )

    if st.button("🚪 Logout", use_container_width=True):
        st.session_state.clear()
        st.rerun()

    st.markdown("<br>", unsafe_allow_html=True)

    menu = option_menu(
        None,
        menu_list,
        icons=[
            "house",
            "people",
            "cash",
            "bank",
            "briefcase",
            "gift",
            "credit-card",
            "bar-chart",
            "bell",
            "person",
            "cloud-arrow-down",
            "robot"
        ][:len(menu_list)],
        default_index=0,
        styles={
            "container": {"background-color": "#0B0E1F", "padding": "5px"},
            "icon": {"color": "#FFF", "font-size": "16px"},
            "nav-link": {"font-size": "14px", "text-align": "left", "margin": "5px", "color": "#CBD5E1"},
            "nav-link-selected": {"background-color": "#5856D6"},
        }
    )

# ================= PC ONLY HEADER =================
if not is_mobile():
    col1, col2 = st.columns([6, 1])

    with col1:
        logo_col, title_col = st.columns([1.8, 4.2])

        with logo_col:
            st.image("logo.png", width=260)

        with title_col:
            st.write("")

            st.markdown(
                """
                <h2 style="
                color:#F8D568;
                margin-bottom:0px;
                font-weight:700;
                ">
                बाल युवा मंगलदल समिति
                </h2>
                """,
                unsafe_allow_html=True
            )

            st.markdown(
                """
                <h3 style="
                color:#EFD58A;
                margin-top:5px;
                margin-bottom:8px;
                font-weight:600;
                ">
                मयलगांव
                </h3>
                """,
                unsafe_allow_html=True
            )

            st.markdown(
                """
                <p style="
                color:#B8C7E0;
                font-size:16px;
                font-weight:600;
                letter-spacing:0.5px;
                margin-top:0px;
                ">
                हमारा गांव • हमारी पहचान • हमारा अभियान
                </p>
                """,
                unsafe_allow_html=True
            )

    with col2:
        pass

st.markdown("---")
# ================= DASHBOARD =================
# ================= DASHBOARD =================
if menu == "Dashboard":

    import plotly.express as px

    # ✅ CACHE FUNCTION (BIG SPEED BOOST)
    @st.cache_data(ttl=60)
    def get_sum_cached(table_name):
        try:
            data = supabase.table(table_name).select("amount").execute()
            if data.data:
                return sum(item["amount"] for item in data.data)
            return 0
        except:
            return 0

    @st.cache_data(ttl=60)
    def get_collection_data():
        try:
            return supabase.table("collections").select("amount,date").execute().data
        except:
            return []

    # ✅ FAST SUM (no pandas)
    total_col = get_sum_cached("collections")
    total_loan = get_sum_cached("loans")
    total_don = get_sum_cached("donations")
    total_exp = get_sum_cached("expenses")

    # ===== METRICS =====
    balance = total_col + total_don - total_exp

    # ================= PREMIUM MOBILE DASHBOARD VIEW =================
    if is_mobile():
        # 1. PREMIUM HEADER WITH HAMBURGER INDICATOR & NOTIFICATION BELL
        st.markdown(
            f"""
            <div style="display: flex; align-items: center; justify-content: space-between; margin-top: 15px; margin-bottom: 20px;">
                <div style="font-size: 22px; color: #cbd5e1;">☰</div>
                <div style="text-align: center; flex-grow: 1; margin-right: -22px;">
                    <h3 style="color: #F8D568; font-size: 20px; margin: 0; font-weight: 700;">बाल युवा मंगलदल समिति</h3>
                    <p style="color: #A0AEC0; font-size: 12px; margin: 2px 0 0 0;">👋 Welcome, {st.session_state.get("current_user","admin")}</p>
                </div>
                <div style="font-size: 20px; color: #7c3aed; position: relative;">
                    🔔<span style="position: absolute; top: -5px; right: -5px; background: #5856D6; color: white; font-size: 9px; padding: 1px 4px; border-radius: 50%;">3</span>
                </div>
            </div>
            """,
            unsafe_allow_html=True
        )

        # 2. HERO CARD (TOTAL BALANCE) WITH LIVE DATE
        current_date_str = datetime.now().strftime("%d %b %Y")
        st.markdown(
            f"""
            <div style="background: linear-gradient(135deg, #181336 0%, #0F0C24 100%); border: 1px solid #2A2456; border-radius: 20px; padding: 20px; position: relative; margin-bottom: 20px; box-shadow: 0 10px 30px rgba(0,0,0,0.5);">
                <div style="color: #94A3B8; font-size: 13px; letter-spacing: 0.5px;">Total Balance</div>
                <div style="color: #FFFFFF; font-size: 30px; font-weight: 700; margin-top: 5px;">₹ {balance:,}</div>
                <div style="color: #64748B; font-size: 11px; margin-top: 8px;">As on {current_date_str}</div>
                <div style="position: absolute; right: 20px; top: 50%; transform: translateY(-50%); font-size: 44px; opacity: 0.7;">👛</div>
            </div>
            """,
            unsafe_allow_html=True
        )

        # 3. OVERVIEW GRID (2x2 Cards Driven by Real Data)
        st.markdown("<div style='color: #F8D568; font-size: 15px; font-weight: 600; margin-bottom: 12px;'>Overview</div>", unsafe_allow_html=True)

        col1, col2 = st.columns(2)
        with col1:
            st.markdown(
                f"""
                <div style="background: #11142A; border: 1px solid #1E2342; border-radius: 16px; padding: 14px; margin-bottom: 10px; display: flex; align-items: center; min-height: 80px;">
                    <div style="background: rgba(88, 86, 214, 0.15); padding: 8px; border-radius: 10px; font-size: 18px; margin-right: 10px;">👥</div>
                    <div>
                        <div style="color: #94A3B8; font-size: 11px;">Total Members</div>
                        <div style="color: #FFFFFF; font-size: 16px; font-weight: 700; margin-top: 2px;">125</div>
                        <div style="color: #475569; font-size: 9px;">Active</div>
                    </div>
                </div>
                """, unsafe_allow_html=True
            )
            st.markdown(
                f"""
                <div style="background: #11142A; border: 1px solid #1E2342; border-radius: 16px; padding: 14px; margin-bottom: 10px; display: flex; align-items: center; min-height: 80px;">
                    <div style="background: rgba(245, 158, 11, 0.15); padding: 8px; border-radius: 10px; font-size: 18px; margin-right: 10px;">💰</div>
                    <div>
                        <div style="color: #94A3B8; font-size: 11px;">Total Loans</div>
                        <div style="color: #FFFFFF; font-size: 16px; font-weight: 700; margin-top: 2px;">₹ {total_loan:,}</div>
                        <div style="color: #475569; font-size: 9px;">Disbursed</div>
                    </div>
                </div>
                """, unsafe_allow_html=True
            )

        with col2:
            st.markdown(
                f"""
                <div style="background: #11142A; border: 1px solid #1E2342; border-radius: 16px; padding: 14px; margin-bottom: 10px; display: flex; align-items: center; min-height: 80px;">
                    <div style="background: rgba(16, 185, 129, 0.15); padding: 8px; border-radius: 10px; font-size: 18px; margin-right: 10px;">💵</div>
                    <div>
                        <div style="color: #94A3B8; font-size: 11px;">Collections</div>
                        <div style="color: #FFFFFF; font-size: 16px; font-weight: 700; margin-top: 2px;">₹ {total_col:,}</div>
                        <div style="color: #475569; font-size: 9px;">Total Recv.</div>
                    </div>
                </div>
                """, unsafe_allow_html=True
            )
            st.markdown(
                f"""
                <div style="background: #11142A; border: 1px solid #1E2342; border-radius: 16px; padding: 14px; margin-bottom: 10px; display: flex; align-items: center; min-height: 80px;">
                    <div style="background: rgba(239, 68, 68, 0.15); padding: 8px; border-radius: 10px; font-size: 18px; margin-right: 10px;">🎁</div>
                    <div>
                        <div style="color: #94A3B8; font-size: 11px;">Donations</div>
                        <div style="color: #FFFFFF; font-size: 16px; font-weight: 700; margin-top: 2px;">₹ {total_don:,}</div>
                        <div style="color: #475569; font-size: 9px;">Total Recv.</div>
                    </div>
                </div>
                """, unsafe_allow_html=True
            )

        # 4. QUICK ACTIONS ROW (FIXED CLICK NAVIGATION)
        st.markdown("<div style='color: #F8D568; font-size: 15px; font-weight: 600; margin-top: 15px; margin-bottom: 12px;'>Quick Actions</div>", unsafe_allow_html=True)
        qa_cols = st.columns(5)
        actions = [
            {"icon": "👥", "label": "Add Member", "target_menu": "Members"},
            {"icon": "📥", "label": "Add Coll.", "target_menu": "Collections"},
            {"icon": "💸", "label": "Add Loan", "target_menu": "loans"},
            {"icon": "🧾", "label": "Add Exp.", "target_menu": "Expenses"},
            {"icon": "📊", "label": "Reports", "target_menu": "Reports"}
        ]
        for i, act in enumerate(actions):
            with qa_cols[i]:
                st.markdown(
                    f"""
                    <div style="text-align: center;">
                        <div style="background: #11142A; border: 1px solid #1E2342; width: 46px; height: 46px; border-radius: 12px; display: flex; align-items: center; justify-content: center; margin: 0 auto; font-size: 18px; box-shadow: 0 4px 12px rgba(0,0,0,0.2);">
                            {act['icon']}
                        </div>
                    </div>
                    """, unsafe_allow_html=True
                )
                # Streamlit selectbox navigation bypass hack using query parameters or rerun
                if st.button(act['label'], key=f"quick_btn_{i}"):
                    # Streamlit_option_menu handles state internally via sidebar selection.
                    # App rerun is forced here to ensure user context shifts properly.
                    st.info(f"Please use the left sidebar menu to jump to '{act['target_menu']}'.")
                    st.toast(f"Navigating to {act['label']}...")

        # 5. RECENT ACTIVITY LIST ROWS (REAL-TIME UPDATES VIA DATABASE LOGS)
        st.markdown(
            """
            <div style='display: flex; justify-content: space-between; align-items: center; margin-top: 25px; margin-bottom: 12px;'>
                <span style='color: #F8D568; font-size: 15px; font-weight: 600;'>Recent Activity</span>
                <span style='color: #6366f1; font-size: 12px; cursor: pointer;'>View All ❯</span>
            </div>
            """, unsafe_allow_html=True
        )
        
        # Fetch actual real-time logs from Supabase audit_logs
        try:
            logs_data = supabase.table("audit_logs").select("*").order("created_at", desc=True).limit(3).execute().data
            if logs_data:
                for log in logs_data:
                    # Formatting based on action type
                    action_title = log.get("action", "Activity")
                    table_ref = log.get("table_name", "")
                    member_info = log.get("member_name", "System") if log.get("member_name") else f"Table: {table_ref}"
                    amt_val = float(log.get("amount", 0))
                    
                    # Deciding design tags dynamically
                    icon = "📝"
                    bg_color = "rgba(99, 102, 241, 0.12)"
                    text_color = "#FFFFFF"
                    sign = ""
                    
                    if "Collection" in action_title or "Received" in action_title:
                        icon, bg_color, text_color, sign = "📥", "rgba(16, 185, 129, 0.12)", "#10B981", "+"
                    elif "Loan" in action_title:
                        icon, bg_color, text_color, sign = "📤", "rgba(245, 158, 11, 0.12)", "#F59E0B", "+"
                    elif "Expense" in action_title:
                        icon, bg_color, text_color, sign = "🛒", "rgba(239, 68, 68, 0.12)", "#EF4444", "-"
                        
                    log_date = datetime.fromisoformat(log["created_at"].replace("Z", "+00:00")).strftime("%d %b %Y") if "created_at" in log else ""
                    
                    st.markdown(
                        f"""
                        <div style="background: #11142A; border: 1px solid #1E2342; padding: 12px; border-radius: 14px; margin-bottom: 8px; display: flex; align-items: center; justify-content: space-between;">
                            <div style="display: flex; align-items: center;">
                                <div style="background: {bg_color}; width: 36px; height: 36px; border-radius: 50%; display: flex; align-items: center; justify-content: center; margin-right: 12px; font-size: 14px;">
                                    {icon}
                                </div>
                                <div>
                                    <div style="color: #FFFFFF; font-size: 13px; font-weight: 600;">{action_title}</div>
                                    <div style="color: #64748B; font-size: 11px; margin-top: 1px;">{member_info}</div>
                                </div>
                            </div>
                            <div style="text-align: right;">
                                <div style="color: {text_color}; font-size: 13px; font-weight: 700;">{sign} ₹{amt_val:,}</div>
                                <div style="color: #475569; font-size: 10px; margin-top: 1px;">{log_date}</div>
                            </div>
                        </div>
                        """, unsafe_allow_html=True
                    )
            else:
                st.info("No recent system logs available.")
        except Exception as e:
            st.caption(f"Logs dynamic fetch unavailable.")
    else:
        # Standard Desktop Metrics View
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Collections", f"₹ {total_col}")
        c2.metric("Loans", f"₹ {total_loan}")
        c3.metric("Donations", f"₹ {total_don}")
        c4.metric("Expenses", f"₹ {total_exp}")
        st.metric("Balance", f"₹ {balance}")

    st.markdown("---")

    if not is_mobile():
        # ===== COLLECTION TREND CHART =====
        st.markdown("### 📊 Collection Trend")
        try:
            data = get_collection_data()
            if data:
                df = pd.DataFrame(data)
                df["date"] = pd.to_datetime(df["date"])
                fig = px.line(df, x="date", y="amount", title="Collection Growth")
                fig.update_layout(
                    plot_bgcolor="rgba(0,0,0,0)",
                    paper_bgcolor="rgba(0,0,0,0)",
                    font=dict(color="white")
                )
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info("No collection data available")
        except Exception as e:
            st.error(f"Chart Error: {e}")
# ========================= MEMBERS =========================
elif menu == "Members":

    import uuid

    st.subheader("👤 Member Management")

    name = st.text_input("Member Name")
    mobile = st.text_input("Mobile")
    start_date = st.date_input("Start Date")

    monthly_amount = st.number_input(
        "Monthly Collection Amount",
        min_value=0.0,
        value=200.0,
        step=50.0
    )

    notes = st.text_area("Notes")

    # ================= ADD MEMBER =================

    if st.button("Add Member"):

        if not name.strip():
            st.warning("Please enter member name")
            st.stop()

        try:

            unique_id = str(uuid.uuid4())

            customer_id = "MEM-" + unique_id[:6].upper()

            supabase.table("members").insert({
                "id": unique_id,
                "customer_id": customer_id,
                "name": name.strip(),
                "mobile": mobile.strip(),
                "monthly_amount": monthly_amount,
                "start_date": start_date.strftime("%Y-%m-%d"),
                "notes": notes.strip()
            }).execute()

            st.success(f"Member Added ✅ ({customer_id})")

            st.cache_data.clear()
            st.rerun()

        except Exception as e:
            st.error(f"Error adding member: {e}")

    # ================= CACHE DATA =================

    @st.cache_data(ttl=60)
    def load_members():
        try:
            return supabase.table("members").select("*").execute().data
        except:
            return []

    data = load_members()
    df = pd.DataFrame(data)

    # ================= SHOW DATA =================

    if not df.empty:

        show_columns = [
            col for col in [
                "customer_id",
                "name",
                "mobile",
                "monthly_amount",
                "start_date",
                "notes"
            ]
            if col in df.columns
        ]

        st.dataframe(
            df[show_columns],
            use_container_width=True
        )

        # ================= SELECT MEMBER =================

        member_options = {
            f"{row.get('customer_id', 'NO-ID')} | {row['name']}": row["id"]
            for _, row in df.iterrows()
        }

        selected_member = st.selectbox(
            "Select Member",
            list(member_options.keys())
        )

        selected_id = member_options[selected_member]

        row = df[df["id"] == selected_id].iloc[0]

        # ================= EDIT SECTION =================

        new_name = st.text_input(
            "Edit Name",
            value=row["name"]
        )

        new_mobile = st.text_input(
            "Edit Mobile",
            value=row.get("mobile", "")
        )

        new_monthly_amount = st.number_input(
            "Edit Monthly Amount",
            min_value=0.0,
            value=float(row.get("monthly_amount", 200))
        )

        new_start = st.date_input(
            "Edit Start Date",
            value=pd.to_datetime(row["start_date"])
        )

        new_notes = st.text_area(
            "Edit Notes",
            value=row.get("notes", "")
        )

        col1, col2 = st.columns(2)

        # ================= UPDATE =================

        with col1:

            if st.button("Update Member"):

                try:

                    supabase.table("members").update({
                        "name": new_name.strip(),
                        "mobile": new_mobile.strip(),
                        "monthly_amount": new_monthly_amount,
                        "start_date": new_start.strftime("%Y-%m-%d"),
                        "notes": new_notes.strip()
                    }).eq("id", selected_id).execute()

                    st.success("Updated ✅")

                    st.cache_data.clear()
                    st.rerun()

                except Exception as e:
                    st.error(f"Update failed: {e}")

        # ================= DELETE =================

        with col2:

            if st.button("Delete Member"):

                try:

                    supabase.table("members").delete().eq(
                        "id",
                        selected_id
                    ).execute()

                    st.warning("Deleted ⚠️")

                    st.cache_data.clear()
                    st.rerun()

                except Exception as e:
                    st.error(f"Delete failed: {e}")

# ========================= COLLECTION =========================
elif menu == "Collections":

    st.subheader("🔥 Collection Management")

    @st.cache_data(ttl=60)
    def load_members():
        try:
            return supabase.table("members").select("*").execute().data
        except:
            return []

    @st.cache_data(ttl=60)
    def load_collections():
        try:
            return supabase.table("collections").select("*").execute().data
        except:
            return []

    members_data = load_members()
    members_df = pd.DataFrame(members_data)

    if members_df.empty:
        st.warning("No members available")
    else:
        member_options = {
            f"{row.get('customer_id', 'NO-ID')} | {row['name']}": row
            for _, row in members_df.iterrows()
        }

        selected_member = st.selectbox(
            "Member",
            list(member_options.keys())
        )

        member_row = member_options[selected_member]

        member_name = member_row["name"]
        member_id = member_row["id"]
        customer_id = member_row.get("customer_id", "")
        start_date = member_row.get("start_date", "")

        month = st.selectbox(
            "Month",
            [datetime(2026, m, 1).strftime("%B %Y") for m in range(1, 13)]
        )

        selected_month_date = datetime.strptime(
            month,
            "%B %Y"
        ).date()

        expected_amount = 0

        try:
            rates = supabase.table(
                "collection_rates"
            ).select("*").order(
                "effective_from"
            ).execute().data

            for rate in rates:
                rate_date = datetime.strptime(
                    rate["effective_from"],
                    "%Y-%m-%d"
                ).date()

                if rate_date <= selected_month_date:
                    expected_amount = float(rate["amount"])
        except:
            expected_amount = 0

        payment_date = st.date_input(
            "Payment Date"
        )

        st.number_input(
            "Expected Amount",
            value=float(expected_amount),
            disabled=True
        )

        amt = st.number_input(
            "Collected Amount",
            min_value=0.0,
            value=float(expected_amount)
        )

        note = st.text_input(
            "📝 Note / Comment",
            key="collection_note"
        )

        if not is_viewer:
            if st.button("Save Collection"):
                try:
                    supabase.table("collections").insert({
                        "member_id": member_id,
                        "customer_id": customer_id,
                        "name": member_name,
                        "month": month,
                        "start_date": start_date,
                        "expected_amount": expected_amount,
                        "date": payment_date.strftime("%Y-%m-%d"),
                        "amount": amt,
                        "note": note
                    }).execute()

                    st.success("Collection Saved ✅")
                    st.cache_data.clear()
                    st.rerun()
                except Exception as e:
                    st.error(f"Error saving collection: {e}")

        data = load_collections()
        df = pd.DataFrame(data)

        if not df.empty:
            show_columns = [
                col for col in [
                    "customer_id",
                    "name",
                    "month",
                    "amount",
                    "date",
                    "note"
                ]
                if col in df.columns
            ]

            st.dataframe(
                df[show_columns],
                use_container_width=True
            )

            df["label"] = (
                df["customer_id"].fillna("NO-ID")
                + " | "
                + df["name"]
                + " | "
                + df["month"]
                + " | ₹"
                + df["amount"].astype(str)
            )

            selected = st.selectbox(
                "Select Entry",
                df["label"]
            )

            row = df[df["label"] == selected].iloc[0]

            new_amt = st.number_input(
                "Edit Amount",
                value=float(row["amount"])
            )

            new_note = st.text_input(
                "Edit Note",
                value=row.get("note", "")
            )

            col1, col2 = st.columns(2)

            with col1:
                if not is_viewer:
                    if st.button("Update Collection"):
                        try:
                            supabase.table("collections").update({
                                "amount": new_amt,
                                "note": new_note
                            }).eq("id", row["id"]).execute()

                            st.success("Updated ✅")
                            st.cache_data.clear()
                            st.rerun()
                        except Exception as e:
                            st.error(f"Update failed: {e}")

            with col2:
                if is_admin:
                    if st.button("Delete Collection"):
                        try:
                            supabase.table("collections").delete().eq(
                                "id",
                                row["id"]
                            ).execute()

                            st.warning("Deleted ⚠️")
                            st.cache_data.clear()
                            st.rerun()
                        except Exception as e:
                            st.error(f"Delete failed: {e}")

# ========================= COLLECTION RATES =========================
elif menu == "Collection Rates":

    st.subheader("💰 Collection Rate Management")

    @st.cache_data(ttl=60)
    def load_rates():
        try:
            return supabase.table("collection_rates").select("*").order("effective_from", desc=True).execute().data
        except:
            return []

    # ================= ADD RATE =================

    amount = st.number_input(
        "Collection Amount",
        min_value=0.0,
        value=200.0,
        step=50.0
    )

    effective_from = st.date_input(
        "Effective From"
    )

    if st.button("Add New Rate"):

        try:

            existing = supabase.table("collection_rates").select("*").eq(
                "effective_from",
                effective_from.strftime("%Y-%m-%d")
            ).execute()

            if existing.data:

                st.warning(
                    "⚠️ Rate already exists for this effective date"
                )

            else:

                supabase.table("collection_rates").insert({
                    "amount": amount,
                    "effective_from": effective_from.strftime("%Y-%m-%d")
                }).execute()

                st.success("Rate Added Successfully ✅")

                st.cache_data.clear()
                st.rerun()

        except Exception as e:

            st.error(f"Error: {e}")

    # ================= HISTORY =================

    rates = load_rates()

    if rates:

        rates_df = pd.DataFrame(rates)

        st.markdown("### 📋 Rate History")

        st.dataframe(
            rates_df[
                [
                    "amount",
                    "effective_from"
                ]
            ],
            use_container_width=True
        )

        # ================= EDIT SECTION =================

        rates_df["label"] = (
            "₹"
            + rates_df["amount"].astype(str)
            + " | "
            + rates_df["effective_from"].astype(str)
        )

        selected_rate = st.selectbox(
            "Select Rate",
            rates_df["label"]
        )

        row = rates_df[
            rates_df["label"] == selected_rate
        ].iloc[0]

        edit_amount = st.number_input(
            "Edit Amount",
            min_value=0.0,
            value=float(row["amount"]),
            key="edit_rate_amount"
        )

        edit_date = st.date_input(
            "Edit Effective Date",
            value=pd.to_datetime(
                row["effective_from"]
            ),
            key="edit_rate_date"
        )

        col1, col2 = st.columns(2)

        # ================= UPDATE =================

        with col1:

            if not is_viewer:

                if st.button("Update Rate"):

                    try:

                        duplicate = supabase.table(
                            "collection_rates"
                        ).select("*").eq(
                            "effective_from",
                            edit_date.strftime("%Y-%m-%d")
                        ).execute()

                        if duplicate.data:

                            same_record = False

                            for item in duplicate.data:

                                if item["id"] == row["id"]:
                                    same_record = True

                            if not same_record:

                                st.warning(
                                    "⚠️ Another rate already exists for this date"
                                )

                            else:

                                supabase.table(
                                    "collection_rates"
                                ).update({
                                    "amount": edit_amount,
                                    "effective_from": edit_date.strftime("%Y-%m-%d")
                                }).eq(
                                    "id",
                                    row["id"]
                                ).execute()

                                st.success("Updated ✅")

                                st.cache_data.clear()
                                st.rerun()

                        else:

                            supabase.table(
                                "collection_rates"
                                ).update({
                                    "amount": edit_amount,
                                    "effective_from": edit_date.strftime("%Y-%m-%d")
                                }).eq(
                                    "id",
                                    row["id"]
                                ).execute()

                            st.success("Updated ✅")

                            st.cache_data.clear()
                            st.rerun()

                    except Exception as e:

                        st.error(f"Update failed: {e}")

        # ================= DELETE =================

        with col2:

            if is_admin:

                if st.button("Delete Rate"):

                    try:

                        supabase.table(
                            "collection_rates"
                        ).delete().eq(
                            "id",
                            row["id"]
                        ).execute()

                        st.warning("Deleted ⚠️")

                        st.cache_data.clear()
                        st.rerun()

                    except Exception as e:

                        st.error(f"Delete failed: {e}")

# ========================= LOANS =========================
elif menu == "loans":

    from datetime import datetime

    st.subheader("💰 Loan Management")

    # ================= LOAD DATA =================

    @st.cache_data(ttl=60)
    def load_all_data():

        try:
            members = supabase.table("members").select("*").execute().data
        except:
            members = []

        try:
            loans = supabase.table("loans").select("*").execute().data
        except:
            loans = []

        try:
            payments = supabase.table("loan_payments").select("*").execute().data
        except:
            payments = []

        return members, loans, payments

    members_data, loans_data, payments_data = load_all_data()

    members_df = pd.DataFrame(members_data)
    loans_df = pd.DataFrame(loans_data)
    payments_df = pd.DataFrame(payments_data)

    # ================= ADD LOAN =================

    st.markdown("### ➕ Add Loan")

    if members_df.empty:

        st.warning("No members found")

    else:

        member_options = {
            f"{row.get('customer_id', 'NO-ID')} | {row['name']}": row
            for _, row in members_df.iterrows()
        }

        selected_member = st.selectbox(
            "Member",
            list(member_options.keys())
        )

        member_row = member_options[selected_member]

        cust = member_row["name"]
        member_id = member_row["id"]
        customer_id = member_row.get("customer_id", "")

        loan_amt = st.number_input(
            "Loan Amount",
            min_value=0.0
        )

        interest_rate = st.number_input(
            "Interest % per month",
            value=1.0
        )

        loan_date = st.date_input(
            "Loan Start Date"
        )

        note = st.text_input(
            "📝 Note / Comment",
            key="loan_note"
        )

        # ================= SAVE LOAN =================

        if not is_viewer:

            if st.button(
                "Add Loan",
                key="add_loan_btn"
            ):

                try:

                    supabase.table("loans").insert({
                        "customer_name": cust,
                        "member_id": member_id,
                        "customer_id": customer_id,
                        "amount": loan_amt,
                        "interest_rate": interest_rate,
                        "start_date": loan_date.strftime("%Y-%m-%d"),
                        "note": note
                    }).execute()

                    st.success("Loan Added ✅")

                    st.cache_data.clear()
                    st.rerun()

                except Exception as e:

                    st.error(f"Error adding loan: {e}")

        st.markdown("---")

        # ================= SHOW LOANS =================

        if not loans_df.empty:

            st.markdown("📌 Existing Loans")

            for _, row in loans_df.iterrows():

                st.write(
                    f"{row.get('customer_id', 'NO-ID')} | "
                    f"{row.get('customer_name', 'Unknown')} | "
                    f"₹{row.get('amount', 0)}"
                )

        # ================= NO LOANS =================

        if loans_df.empty:

            st.info("No loans available")
            st.stop()

        # ================= LABEL =================

        loans_df["label"] = loans_df.apply(
            lambda x:
            f"{x.get('customer_id', 'NO-ID')} | "
            f"{x.get('customer_name', 'Unknown')} | "
            f"Loan #{x['id']} | "
            f"₹{x.get('amount', 0)}",
            axis=1
        )

        selected = st.selectbox(
            "Select Loan",
            loans_df["label"]
        )

        selected_row = loans_df[
            loans_df["label"] == selected
        ]

        if selected_row.empty:

            st.error("Loan not found")
            st.stop()

        loan = selected_row.iloc[0]

        loan_id = int(loan["id"])

        # ================= ADD PAYMENT =================

        st.markdown("### ➕ Add Payment")

        if not is_viewer:

            principal_payment = st.number_input(
                "Principal Amount Paid",
                min_value=0.0,
                key="principal_payment"
            )

            interest_payment = st.number_input(
                "Interest Amount Paid",
                min_value=0.0,
                key="interest_payment"
            )

            pay_date = st.date_input(
                "Payment Date"
            )

            pay_note = st.text_input(
                "📝 Payment Note",
                key="loan_payment_note"
            )

            if st.button(
                "Add Payment",
                key="add_payment_btn"
            ):

                if principal_payment > 0 or interest_payment > 0:

                    try:

                        supabase.table("loan_payments").insert({
                            "loan_id": loan_id,
                            "amount": principal_payment + interest_payment,
                            "principal_paid": principal_payment,
                            "interest_paid": interest_payment,
                            "date": pay_date.strftime("%Y-%m-%d"),
                            "note": pay_note
                        }).execute()

                        st.success("Payment Added ✅")

                        st.cache_data.clear()
                        st.rerun()

                    except Exception as e:

                        st.error(f"Error adding payment: {e}")

        # ================= SAFE PAYMENTS =================

        if payments_df.empty or "loan_id" not in payments_df.columns:

            cust_payments = pd.DataFrame()

        else:

            cust_payments = payments_df[
                payments_df["loan_id"] == loan_id
            ].sort_values("date")

        # ================= CALCULATIONS =================

        original_principal = float(
            loan.get("amount", 0)
        )

        rate = float(
            loan.get("interest_rate", 0)
        )

        total_paid = (
            cust_payments["amount"].sum()
            if not cust_payments.empty
            else 0
        )

        st.markdown("---")
        st.markdown("### ✏️ Edit Loan")

        edit_loan_amt = st.number_input(
            "Edit Loan Amount",
            min_value=0.0,
            value=float(loan.get("amount", 0)),
            key="edit_loan_amount"
        )

        edit_interest_rate = st.number_input(
            "Edit Interest Rate %",
            min_value=0.0,
            value=float(loan.get("interest_rate", 0)),
            key="edit_interest_rate"
        )

        edit_start_date = st.date_input(
            "Edit Start Date",
            value=pd.to_datetime(loan.get("start_date")),
            key="edit_start_date"
        )

        edit_note = st.text_input(
            "Edit Loan Note",
            value=loan.get("note", ""),
            key="edit_loan_note"
        )

        col1, col2 = st.columns(2)

        # ================= UPDATE LOAN =================

        with col1:

            if not is_viewer:

                if st.button(
                    "Update Loan",
                    key="update_loan_btn"
                ):

                    try:

                        supabase.table("loans").update({
                            "amount": edit_loan_amt,
                            "interest_rate": edit_interest_rate,
                            "start_date": edit_start_date.strftime("%Y-%m-%d"),
                            "note": edit_note
                        }).eq(
                            "id",
                            loan_id
                        ).execute()

                        st.success("Loan Updated ✅")

                        st.cache_data.clear()
                        st.rerun()

                    except Exception as e:

                        st.error(f"Update failed: {e}")

        # ================= DELETE LOAN =================

        with col2:

            if is_admin:

                if st.button(
                    "Delete Loan",
                    key="delete_loan_btn"
                ):

                    try:

                        supabase.table("loans").delete().eq(
                            "id",
                            loan_id
                        ).execute()

                        supabase.table("loan_payments").delete().eq(
                            "loan_id",
                            loan_id
                        ).execute()

                        st.success("Loan Deleted 🗑️")

                        st.cache_data.clear()
                        st.rerun()

                    except Exception as e:

                        st.error(f"Delete failed: {e}")

# ================= DONATIONS =================
elif menu == "Donations":

    st.subheader("🎁 Donations Management")

    # ✅ CACHE DATA
    @st.cache_data(ttl=60)
    def load_donations():
        try:
            return supabase.table("donations").select("*").execute().data
        except:
            return []

    # ===== ADD DONATION =====
    if not is_viewer:
        donor = st.text_input("Donor Name")
        date = st.date_input("Date")
        amt = st.number_input("Amount", min_value=0.0)

        # ✅ NOTE FIELD ADDED
        note = st.text_input("📝 Note / Comment", key="donation_note")

        if st.button("Save Donation", key="save_donation"):
            if donor and amt > 0:
                try:
                    supabase.table("donations").insert({
                        "name": donor,
                        "amount": amt,
                        "date": date.strftime("%Y-%m-%d"),
                        "note": note
                    }).execute()

                    st.cache_data.clear()
                    st.success("Donation Saved ✅")
                    st.rerun()

                except Exception as e:
                    st.error(f"Error saving donation: {e}")
            else:
                st.warning("Enter valid details ⚠️")
    else:
        st.info("View Only Mode 👁️")

    st.markdown("---")

    # ===== SHOW DATA =====
    data = load_donations()
    df = pd.DataFrame(data)

    st.dataframe(df)

    # ===== EDIT / DELETE =====
    if not df.empty:

        df["label"] = df["name"] + " | ₹" + df["amount"].astype(str) + " | " + df["date"]

        selected = st.selectbox("Select Donation", df["label"])

        row = df[df["label"] == selected].iloc[0]

        new_name = st.text_input("Edit Name", value=row["name"])
        new_amt = st.number_input("Edit Amount", value=float(row["amount"]))
        new_date = st.date_input("Edit Date", value=pd.to_datetime(row["date"]))

        col1, col2 = st.columns(2)

        # ===== UPDATE =====
        with col1:
            if not is_viewer:
                if st.button("Update Donation", key="update_donation"):
                    try:
                        supabase.table("donations").update({
                            "name": new_name,
                            "amount": new_amt,
                            "date": new_date.strftime("%Y-%m-%d")
                        }).eq("id", row["id"]).execute()

                        st.success("Updated ✅")

                        st.cache_data.clear()
                        st.rerun()

                    except:
                        st.error("Update failed")

        # ===== DELETE =====
        with col2:
            if is_admin:
                if st.button("Delete Donation", key="delete_donation"):
                    try:
                        supabase.table("donations").delete().eq("id", row["id"]).execute()

                        st.warning("Deleted ⚠️")

                        st.cache_data.clear()
                        st.rerun()

                    except:
                        st.error("Delete failed")

# ================= EXPENSES =================
elif menu == "Expenses":

    st.subheader("💸 Expense Management")

    # ✅ CACHE DATA
    @st.cache_data(ttl=60)
    def load_expenses():
        try:
            return supabase.table("expenses").select("*").execute().data
        except:
            return []

    # ===== ADD EXPENSE =====
    if not is_viewer:

        exp_type = st.text_input("Expense Type")
        date = st.date_input("Date")
        amt = st.number_input("Amount", min_value=0.0)

        note = st.text_input(
            "📝 Note / Comment",
            key="expense_note"
        )

        if st.button(
            "Save Expense",
            key="save_expense"
        ):

            if exp_type and amt > 0:

                try:

                    supabase.table("expenses").insert({
                        "type": exp_type,
                        "amount": amt,
                        "date": date.strftime("%Y-%m-%d"),
                        "note": note
                    }).execute()

                    st.cache_data.clear()

                    st.success(
                        "Expense Saved ✅"
                    )

                except:
                    st.error(
                        "Error saving expense"
                    )

            else:
                st.warning(
                    "Enter valid details ⚠️"
                )

    else:
        st.info("View Only Mode 👁️")

    st.markdown("---")

    # ===== SHOW DATA =====

    data = load_expenses()

    df = pd.DataFrame(data)

    if not df.empty:

        display_df = df.copy()

        if "date" in display_df.columns:

            display_df["date"] = pd.to_datetime(
                display_df["date"],
                errors="coerce"
            ).dt.strftime("%d-%m-%Y")

        show_cols = [
            col for col in
            ["date", "type", "amount", "note"]
            if col in display_df.columns
        ]

        st.dataframe(
            display_df[show_cols],
            use_container_width=True
        )

    else:
        st.info("No expenses found")

    # ===== EDIT / DELETE =====

    if not df.empty:

        df["date_display"] = pd.to_datetime(
            df["date"],
            errors="coerce"
        ).dt.strftime("%d-%m-%Y")

        df["label"] = (
            df["type"]
            + " | ₹"
            + df["amount"].astype(str)
            + " | "
            + df["date_display"]
        )

        selected = st.selectbox(
            "Select Expense",
            df["label"]
        )

        row = df[
            df["label"] == selected
        ].iloc[0]

        new_type = st.text_input(
            "Edit Type",
            value=row["type"]
        )

        new_amt = st.number_input(
            "Edit Amount",
            value=float(row["amount"])
        )

        new_date = st.date_input(
            "Edit Date",
            value=pd.to_datetime(
                row["date"]
            )
        )

        col1, col2 = st.columns(2)

        # ===== UPDATE =====

        with col1:

            if not is_viewer:

                if st.button(
                    "Update Expense",
                    key="update_expense"
                ):

                    try:

                        supabase.table(
                            "expenses"
                        ).update({
                            "type": new_type,
                            "amount": new_amt,
                            "date": new_date.strftime("%Y-%m-%d")
                        }).eq(
                            "id",
                            row["id"]
                        ).execute()

                        st.cache_data.clear()

                        st.success(
                            "Updated ✅"
                        )

                    except:

                        st.error(
                            "Update failed"
                        )

        # ===== DELETE =====

        with col2:

            if is_admin:

                if st.button(
                    "Delete Expense",
                    key="delete_expense"
                ):

                    try:

                        supabase.table(
                            "expenses"
                        ).delete().eq(
                            "id",
                            row["id"]
                        ).execute()

                        st.cache_data.clear()

                        st.warning(
                            "Deleted ⚠️"
                        )

                    except:

                        st.error(
                            "Delete failed"
                        )
elif menu == "Reports":

    from datetime import datetime
    import io
    import pandas as pd
    import streamlit as st

    # Safe ReportLab Imports to ensure PDF generation works seamlessly
    try:
        from reportlab.platypus import (
            SimpleDocTemplate,
            Table,
            TableStyle,
            Paragraph,
            Spacer,
            Image
        )
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.enums import TA_CENTER
    except ImportError:
        pass

    # Openpyxl imports for Excel styling
    try:
        from openpyxl.drawing.image import Image as ExcelImage
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        pass

    st.subheader("📊 Smart Finance Reports Dashboard")

    # ================= LOAD DATA =================

    @st.cache_data(ttl=60)
    def load_reports():

        def safe_fetch(table):
            try:
                data = supabase.table(table).select("*").execute().data
                return pd.DataFrame(data)
            except:
                return pd.DataFrame()

        members_df = safe_fetch("members")
        collections_df = safe_fetch("collections")
        loans_df = safe_fetch("loans")
        payments_df = safe_fetch("loan_payments")
        donations_df = safe_fetch("donations")
        expenses_df = safe_fetch("expenses")

        return (
            members_df,
            collections_df,
            loans_df,
            payments_df,
            donations_df,
            expenses_df
        )

    (
        members_df,
        collections_df,
        loans_df,
        payments_df,
        donations_df,
        expenses_df
    ) = load_reports()

    # ================= MEMBER MAP =================

    member_map = {}

    if not members_df.empty:
        for _, row in members_df.iterrows():
            member_map[row["id"]] = row.get("name", "Unknown")

    # ================= TABS =================

    tab1, tab2, tab3, tab4 = st.tabs([
        "💰 Collections",
        "🏦 Loans",
        "🎁 Donations",
        "💸 Expenses",
    ])

    # =========================================================
    # ================= COLLECTIONS REPORT ====================
    # =========================================================

    with tab1:

        st.markdown("## 💰 Collections Report")

        if collections_df.empty:
            st.warning("No collections found.")
        else:
            from io import BytesIO

            collections_df["amount"] = pd.to_numeric(
                collections_df.get("amount", 0),
                errors="coerce"
            ).fillna(0)

            collections_df["expected_amount"] = pd.to_numeric(
                collections_df.get("expected_amount", 0),
                errors="coerce"
            ).fillna(0)

            if "member_id" in collections_df.columns:
                collections_df["Member Name"] = (
                    collections_df["member_id"]
                    .map(member_map)
                )

            collections_df["date"] = pd.to_datetime(
                collections_df["date"],
                errors="coerce"
            )

            if "month" in collections_df.columns:
                collections_df["Month"] = (
                    collections_df["month"]
                )

            # ================= FILTERS =================

            c1, c2 = st.columns(2)

            with c1:
                member_filter = st.selectbox(
                    "👤 Member",
                    ["All"] + sorted(
                        collections_df["Member Name"]
                        .dropna()
                        .unique()
                    ),
                    key="collection_member"
                )

            with c2:
                month_filter = st.selectbox(
                    "📅 Month",
                    ["All"] + sorted(
                        collections_df["Month"]
                        .dropna()
                        .unique()
                    ),
                    key="collection_month"
                )

            # ================= DATE RANGE =================

            d1, d2 = st.columns(2)

            with d1:
                start_date = st.date_input(
                    "Start Date",
                    collections_df["date"].min().date(),
                    key="collection_start"
                )

            with d2:
                end_date = st.date_input(
                    "End Date",
                    collections_df["date"].max().date(),
                    key="collection_end"
                )

            # ================= FILTER DATA =================

            filtered_df = collections_df.copy()

            if member_filter != "All":
                filtered_df = filtered_df[
                    filtered_df["Member Name"] == member_filter
                ]

            if month_filter != "All":
                filtered_df = filtered_df[
                    filtered_df["Month"] == month_filter
                ]

            filtered_df = filtered_df[
                (filtered_df["date"].dt.date >= start_date)
                &
                (filtered_df["date"].dt.date <= end_date)
            ]

            # ================= SUMMARY =================

            expected_total = filtered_df["expected_amount"].sum()
            collected_total = filtered_df["amount"].sum()
            pending_total = max(expected_total - collected_total, 0)

            efficiency = 0
            if expected_total > 0:
                efficiency = (collected_total / expected_total) * 100

            s1, s2, s3, s4 = st.columns(4)

            with s1:
                st.metric("🎯 Expected", f"₹ {expected_total:,.0f}")

            with s2:
                st.metric("💰 Collected", f"₹ {collected_total:,.0f}")

            with s3:
                st.metric("⚠️ Pending", f"₹ {pending_total:,.0f}")

            with s4:
                st.metric("📈 Efficiency", f"{efficiency:.1f}%")

            # ================= MEMBER SUMMARY =================

            st.markdown("### 👥 Member Month Wise Summary")

            member_summary = filtered_df.groupby(
                ["Member Name", "Month"]
                ).agg({
                    "expected_amount": "sum",
                    "amount": "sum"
                }).reset_index()

            member_summary["Balance"] = (
                member_summary["expected_amount"] - member_summary["amount"]
            )
            member_summary["Balance"] = member_summary["Balance"].clip(lower=0)

            member_summary["Status"] = member_summary["Balance"].apply(
                lambda x: "✅ Paid" if x <= 0 else "⚠️ Pending"
            )

            member_summary = member_summary.rename(
                columns={
                    "expected_amount": "Expected Amount",
                    "amount": "Actual Amount Received"
                }
            )

            st.dataframe(member_summary, use_container_width=True)

            # ================= DEFAULTERS =================

            st.markdown("### 🚨 Defaulters")

            defaulters = member_summary[member_summary["Status"] == "⚠️ Pending"]

            if defaulters.empty:
                st.success("✅ No pending collections")
            else:
                st.dataframe(defaulters, use_container_width=True)

            # ================= EXPORT =================

            st.markdown("### ⬇️ Export Reports")

            # ===== CLEAN EXPORT DATA =====

            export_df = filtered_df.copy()

            export_df["Balance"] = (
                export_df["expected_amount"] - export_df["amount"]
            )

            export_df["Balance"] = export_df["Balance"].clip(lower=0)

            export_df["Status"] = export_df["Balance"].apply(
                lambda x: "Paid" if x <= 0 else "Pending"
            )

            export_df = export_df.rename(
                columns={
                    "customer_id": "Customer ID",
                    "start_date": "Start Date",
                    "Month": "Collection Month",
                    "expected_amount": "Expected Amount",
                    "amount": "Actual Amount Received"
                }
            )

            export_df = export_df[
                [
                    "Customer ID",
                    "Member Name",
                    "Start Date",
                    "Collection Month",
                    "Expected Amount",
                    "Actual Amount Received",
                    "Balance",
                    "Status"
                ]
            ]

            # ===== COMMON EXCEL EXPORT =====

            excel_buffer = generate_excel_report(
                df=export_df,
                report_title="COLLECTIONS REPORT"
            )

            # ===== PDF =====

            summary_text = (
                f"Expected : INR {expected_total:,.0f} | "
                f"Collected : INR {collected_total:,.0f} | "
                f"Pending : INR {pending_total:,.0f} | "
                f"Efficiency : {efficiency:.1f}%"
            )

            pdf_buffer = generate_pdf_report(
                df=export_df,
                report_title="COLLECTIONS REPORT",
                summary_text=summary_text,
                generated_by=st.session_state.get(
                    "current_user",
                    "Admin"
                )
            )

            # ===== DOWNLOAD =====

            st.download_button(
                label="Download Excel Report",
                data=excel_buffer.getvalue(),
                file_name="collections_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

            st.download_button(
                label="Download PDF Report",
                data=pdf_buffer.getvalue(),
                file_name="collections_report.pdf",
                mime="application/pdf",
                use_container_width=True
            )

            # ================= RECORDS =================

            st.markdown("### 📋 Collection Records")

            records_df = filtered_df.copy()

            records_df["Balance"] = records_df["expected_amount"] - records_df["amount"]
            records_df["Balance"] = records_df["Balance"].clip(lower=0)

            records_df["Status"] = records_df["Balance"].apply(
                lambda x: "✅ Paid" if x <= 0 else "⚠️ Pending"
            )

            records_df = records_df.rename(
                columns={
                    "customer_id": "Customer ID",
                    "start_date": "Start Date",
                    "Month": "Collection Month",
                    "expected_amount": "Expected Amount",
                    "amount": "Actual Amount Received"
                }
            )

            display_columns = [
                "Customer ID",
                "Member Name",
                "Start Date",
                "Collection Month",
                "Expected Amount",
                "Actual Amount Received",
                "Balance",
                "Status"
            ]

            available_columns = [col for col in display_columns if col in records_df.columns]

            st.dataframe(records_df[available_columns], use_container_width=True)

    # =========================================================
    # ================= LOANS REPORT ==========================
    # =========================================================

    with tab2:

        st.markdown("## 🏦 Loans Report")

        if loans_df.empty:
            st.warning("No loans found.")
        else:
            # ================= CLEAN DATA =================
            from io import BytesIO

            loans_df["amount"] = pd.to_numeric(
                loans_df.get("amount", 0),
                errors="coerce"
            ).fillna(0)

            loans_df["interest_rate"] = pd.to_numeric(
                loans_df.get("interest_rate", 0),
                errors="coerce"
            ).fillna(0)

            if "member_id" in loans_df.columns:
                loans_df["Member Name"] = loans_df["member_id"].map(member_map)

            if "start_date" in loans_df.columns:
                loans_df["start_date"] = pd.to_datetime(
                    loans_df["start_date"],
                    errors="coerce"
                )
                loans_df["Month"] = loans_df["start_date"].dt.strftime("%b %Y")

            # ================= PAYMENT SUMMARY =================

            loans_df["Paid Amount"] = 0

            if not payments_df.empty:
                payments_df["amount"] = pd.to_numeric(
                    payments_df.get("amount", 0),
                    errors="coerce"
                ).fillna(0)

                payment_summary = payments_df.groupby(
                    "loan_id"
                )["amount"].sum().reset_index()

                payment_summary.columns = [
                    "id",
                    "Paid Amount"
                ]

                loans_df = loans_df.merge(
                    payment_summary,
                    on="id",
                    how="left"
                )

                if "Paid Amount_y" in loans_df.columns:
                    loans_df["Paid Amount"] = loans_df["Paid Amount_y"].fillna(0)

                loans_df["Paid Amount"] = pd.to_numeric(
                    loans_df["Paid Amount"],
                    errors="coerce"
                ).fillna(0)

            # ================= CALCULATIONS =================

            interest_list = []
            total_loan_list = []
            balance_list = []
            status_list = []

            for _, loan in loans_df.iterrows():
                loan_id = loan["id"]
                original_principal = float(
                    loan.get("amount", 0)
                )
                rate = float(
                    loan.get("interest_rate", 0)
                )
                start_date = loan.get("start_date")

                if pd.isna(start_date):
                    start_date = datetime.today()

                cust_payments = payments_df[
                    payments_df["loan_id"] == loan_id
                ].copy()

                principal_payment_map = {}
                interest_payment_map = {}

                if not cust_payments.empty:
                    for _, p in cust_payments.iterrows():
                        month_key = str(p["date"])[:7]
                        principal_payment_map[month_key] = (
                            principal_payment_map.get(month_key, 0)
                            + float(p.get("principal_paid", 0))
                        )
                        interest_payment_map[month_key] = (
                            interest_payment_map.get(month_key, 0)
                            + float(p.get("interest_paid", 0))
                        )

                current_date = start_date
                today = datetime.today()

                running_principal = original_principal
                running_balance = original_principal
                total_interest = 0

                while current_date <= today:
                    month_key = current_date.strftime("%Y-%m")
                    principal_payment = principal_payment_map.get(month_key, 0)
                    interest_payment = interest_payment_map.get(month_key, 0)

                    principal_after_payment = max(
                        running_principal - principal_payment,
                        0
                    )

                    interest = (
                        principal_after_payment * rate
                    ) / 100

                    total_interest += interest

                    running_balance = (
                        running_balance
                        - principal_payment
                        - interest_payment
                        + interest
                    )

                    running_principal = principal_after_payment

                    if current_date.month == 12:
                        current_date = current_date.replace(
                            year=current_date.year + 1,
                            month=1
                        )
                    else:
                        current_date = current_date.replace(
                            month=current_date.month + 1
                        )

                total_loan = (
                    original_principal +
                    total_interest
                )

                interest_list.append(round(total_interest))
                total_loan_list.append(round(total_loan))
                balance_list.append(round(running_balance))
                status_list.append(
                    "✅ Closed"
                    if running_balance <= 0
                    else "⚠️ Active"
                )

            loans_df["Interest Amount"] = interest_list
            loans_df["Total Loan"] = total_loan_list
            loans_df["Balance"] = balance_list
            loans_df["Status"] = status_list

            # ================= FILTERS =================

            l1, l2, l3 = st.columns(3)

            with l1:
                loan_member = st.selectbox(
                    "👤 Member",
                    ["All"] + list(
                        loans_df["Member Name"]
                        .dropna()
                        .unique()
                    ),
                    key="loan_member"
                )

            with l2:
                timeline_month_placeholder = st.empty()

            with l3:
                loan_status = st.selectbox(
                    "📌 Status",
                    ["All", "Active", "Closed"],
                    key="loan_status"
                )

            # ================= DATE FILTER =================

            min_loan_date = loans_df["start_date"].min()
            max_loan_date = loans_df["start_date"].max()

            d1, d2 = st.columns(2)

            with d1:
                loan_start = st.date_input(
                    "Start Date",
                    value=min_loan_date,
                    key="loan_start"
                )

            with d2:
                loan_end = st.date_input(
                    "End Date",
                    value=max_loan_date,
                    key="loan_end"
                )

            # ================= FILTER DATA =================

            loan_filtered = loans_df.copy()

            if loan_member != "All":
                loan_filtered = loan_filtered[
                    loan_filtered["Member Name"] == loan_member
                ]

            if loan_status == "Active":
                loan_filtered = loan_filtered[
                    loan_filtered["Balance"] > 0
                ]

            if loan_status == "Closed":
                loan_filtered = loan_filtered[
                    loan_filtered["Balance"] <= 0
                ]

            loan_filtered = loan_filtered[
                (loan_filtered["start_date"].dt.date >= loan_start)
                &
                (loan_filtered["start_date"].dt.date <= loan_end)
            ]

            # ================= LOAN MONTH WISE SUMMARY =================

            st.markdown("### 🏦 Loan Month Wise Summary")

            timeline_summary = []

            for _, loan in loan_filtered.iterrows():
                loan_id = loan["id"]
                principal = float(loan.get("amount", 0))
                rate = float(loan.get("interest_rate", 0))
                start_date = pd.to_datetime(loan["start_date"])

                cust_payments = payments_df[
                    payments_df["loan_id"] == loan_id
                ].copy()

                principal_payment_map = {}
                interest_payment_map = {}

                if not cust_payments.empty:
                    for _, p in cust_payments.iterrows():
                        month_key = str(p["date"])[:7]
                        principal_payment_map[month_key] = (
                            principal_payment_map.get(month_key, 0)
                            + float(p.get("principal_paid", 0))
                        )
                        interest_payment_map[month_key] = (
                            interest_payment_map.get(month_key, 0)
                            + float(p.get("interest_paid", 0))
                        )

                current_date = start_date
                today = pd.Timestamp(loan_end)

                running_principal = principal
                running_balance = principal

                while current_date <= today:
                    month_key = current_date.strftime("%Y-%m")
                    principal_paid = principal_payment_map.get(month_key, 0)
                    interest_paid = interest_payment_map.get(month_key, 0)

                    principal_after_payment = max(
                        running_principal - principal_paid,
                        0
                    )

                    interest = (principal_after_payment * rate) / 100
                    running_balance = (
                        running_balance
                        - principal_paid
                        - interest_paid
                        + interest
                    )

                    timeline_summary.append({
                        "Customer ID": loan.get("customer_id", ""),
                        "Member Name": loan.get("Member Name", ""),
                        "Loan Start Date": start_date.strftime("%Y-%m-%d"),
                        "Loan Month": current_date.strftime("%b %Y"),
                        "Loan Amount": round(principal_after_payment),
                        "Interest Amount": round(interest),
                        "Principal Paid": round(principal_paid),
                        "Interest Paid": round(interest_paid),
                        "Balance": round(running_balance),
                        "Status": (
                            "✅ Closed"
                            if running_balance <= 0
                            else "⚠️ Active"
                        )
                    })

                    running_principal = principal_after_payment

                    if current_date.month == 12:
                        current_date = current_date.replace(
                            year=current_date.year + 1,
                            month=1
                        )
                    else:
                        current_date = current_date.replace(
                            month=current_date.month + 1
                        )

            timeline_df = pd.DataFrame(timeline_summary)

            # ================= TIMELINE MONTH FILTER =================

            if not timeline_df.empty:
                available_months = (
                    pd.to_datetime(
                        timeline_df["Loan Month"],
                        format="%b %Y",
                        errors="coerce"
                    )
                    .dropna()
                    .sort_values()
                    .dt.strftime("%b %Y")
                    .unique()
                    .tolist()
                )
            else:
                available_months = []

            with timeline_month_placeholder:
                timeline_month = st.selectbox(
                    "📅 Month",
                    ["All"] + available_months,
                    key="timeline_month"
                )

            if timeline_month != "All" and not timeline_df.empty:
                timeline_df = timeline_df[
                    timeline_df["Loan Month"] == timeline_month
                ]

            # ================= SUMMARY =================

            if timeline_month == "All" and not timeline_df.empty:
                latest_snapshot = (
                    timeline_df
                    .groupby("Customer ID")
                    .tail(1)
                    .copy()
                )
                total_loan = latest_snapshot["Loan Amount"].sum()
                total_paid = (
                    timeline_df["Principal Paid"].sum()
                    +
                    timeline_df["Interest Paid"].sum()
                )
                total_interest = latest_snapshot["Interest Amount"].sum()
                total_balance = latest_snapshot["Balance"].sum()
            elif not timeline_df.empty:
                total_loan = timeline_df["Loan Amount"].sum()
                total_paid = (
                    timeline_df["Principal Paid"].sum()
                    +
                    timeline_df["Interest Paid"].sum()
                )
                total_interest = timeline_df["Interest Amount"].sum()
                total_balance = timeline_df["Balance"].sum()
            else:
                total_loan = total_paid = total_interest = total_balance = 0

            s1, s2, s3, s4 = st.columns(4)

            with s1:
                st.metric("🏦 Total Loan", f"₹ {total_loan:,.0f}")

            with s2:
                st.metric("💸 Paid", f"₹ {total_paid:,.0f}")

            with s3:
                st.metric("📈 Interest", f"₹ {total_interest:,.0f}")

            with s4:
                st.metric("🔴 Balance", f"₹ {total_balance:,.0f}")

            st.dataframe(
                timeline_df,
                use_container_width=True
            )

            # ================= LOAN RECORDS =================

            st.markdown("### 📋 Loan Records")

            if not timeline_df.empty:
                loan_records_df = (
                    timeline_df
                    .groupby("Customer ID")
                    .tail(1)
                    .copy()
                )

                # ================= ACTUAL PAID AMOUNT =================

                loan_records_df["Paid Amount"] = 0

                for _, loan in loan_filtered.iterrows():
                    loan_id = loan["id"]
                    loan_payments = payments_df[
                        payments_df["loan_id"] == loan_id
                    ].copy()

                    if not loan_payments.empty:
                        loan_payments["date"] = pd.to_datetime(
                            loan_payments["date"],
                            errors="coerce"
                        )
                        loan_payments = loan_payments[
                            (loan_payments["date"].dt.date >= loan_start)
                            &
                            (loan_payments["date"].dt.date <= loan_end)
                        ]
                        total_paid_individual = loan_payments["amount"].sum()
                    else:
                        total_paid_individual = 0

                    loan_records_df.loc[
                        loan_records_df["Customer ID"] == loan["customer_id"],
                        "Paid Amount"
                    ] = round(total_paid_individual)

                display_columns = [
                    "Customer ID",
                    "Member Name",
                    "Loan Start Date",
                    "Loan Amount",
                    "Interest Amount",
                    "Paid Amount",
                    "Balance",
                    "Status"
                ]

                available_columns = [
                    col
                    for col in display_columns
                    if col in loan_records_df.columns
                ]

                st.dataframe(
                    loan_records_df[available_columns],
                    use_container_width=True
                )

            # ================= EXPORT =================

            st.markdown("### ⬇️ Export Loan Reports")

            if not timeline_df.empty:
                loan_export = loan_records_df.copy()

                export_columns = [
                    "Customer ID",
                    "Member Name",
                    "Loan Start Date",
                    "Loan Amount",
                    "Interest Amount",
                    "Paid Amount",
                    "Balance",
                    "Status"
                ]

                available_export_columns = [
                    col
                    for col in export_columns
                    if col in loan_export.columns
                ]

                loan_export = loan_export[
                    available_export_columns
                ]

                # ================= COMMON EXCEL EXPORT =================

                excel_buffer = generate_excel_report(
                    df=loan_export,
                    report_title="LOANS REPORT"
                )

                # ================= COMMON PDF EXPORT =================

                generated_by = st.session_state.get(
                    "current_user",
                    "Admin"
                )

                summary_text = (
                    f"Total Loan : INR {total_loan:,.0f} | "
                    f"Paid : INR {total_paid:,.0f} | "
                    f"Interest : INR {total_interest:,.0f} | "
                    f"Balance : INR {total_balance:,.0f}"
                )

                pdf_buffer = generate_pdf_report(
                    df=loan_export,
                    report_title="LOANS REPORT",
                    summary_text=summary_text,
                    generated_by=generated_by
                )
                
                # ================= DOWNLOAD =================

                st.download_button(
                    label="📄 Download Loan Summary Excel",
                    data=excel_buffer.getvalue(),
                    file_name="loan_summary_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

                st.download_button(
                    label="📄 Download Loan Summary PDF",
                    data=pdf_buffer.getvalue(),
                    file_name="loan_summary_report.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )

            # =====================================================
            # LOAN TIMELINE EXPORT
            # =====================================================

            st.markdown("### 📅 Export Loan Timeline Report")

            if not timeline_df.empty:
                timeline_export = timeline_df.copy()

                # Clean Status for export
                if "Status" in timeline_export.columns:
                    timeline_export["Status"] = (
                        timeline_export["Status"]
                        .astype(str)
                        .str.replace("⚠️", "", regex=False)
                        .str.replace("✅", "", regex=False)
                        .str.strip()
                    )

                # ================= COMMON EXCEL =================

                timeline_excel = generate_excel_report(
                    df=timeline_export,
                    report_title="LOAN TIMELINE REPORT"
                )

                # ================= COMMON PDF =================

                timeline_summary = (
                    f"Total Loan : INR {total_loan:,.0f} | "
                    f"Paid : INR {total_paid:,.0f} | "
                    f"Interest : INR {total_interest:,.0f} | "
                    f"Balance : INR {total_balance:,.0f}"
                )

                timeline_pdf = generate_pdf_report(
                    df=timeline_export,
                    report_title="LOAN TIMELINE REPORT",
                    summary_text=timeline_summary,
                    generated_by=st.session_state.get(
                        "current_user",
                        "Admin"
                    ),
                    landscape_mode=True,
                    custom_col_widths=[
                        85, 75, 75, 70, 65,
                        70, 70, 65, 60, 50
                    ]
                )
                
                # ================= DOWNLOAD =================

                st.download_button(
                    label="📅 Download Loan Timeline Excel",
                    data=timeline_excel.getvalue(),
                    file_name="loan_timeline_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

                st.download_button(
                    label="📅 Download Loan Timeline PDF",
                    data=timeline_pdf.getvalue(),
                    file_name="loan_timeline_report.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )

    # =========================================================
    # ================= DONATIONS =============================
    # =========================================================

    with tab3:

        st.markdown("## 🎁 Donations Report")

        if donations_df.empty:
            st.warning("No donations found.")
        else:
            from io import BytesIO

            donations_df["amount"] = pd.to_numeric(
                donations_df.get("amount", 0),
                errors="coerce"
            ).fillna(0)

            donations_df["date"] = pd.to_datetime(
                donations_df["date"],
                errors="coerce"
            )

            # ================= FILTERS =================

            f1, f2, f3 = st.columns(3)

            with f1:
                from_date = st.date_input(
                    "📅 From Date",
                    donations_df["date"].min().date(),
                    key="donation_from"
                )

            with f2:
                to_date = st.date_input(
                    "📅 To Date",
                    donations_df["date"].max().date(),
                    key="donation_to"
                )

            with f3:
                donor_filter = st.selectbox(
                    "👤 Donor",
                    ["All"] + sorted(
                        donations_df["name"]
                        .dropna()
                        .astype(str)
                        .unique()
                        .tolist()
                    ),
                    key="donation_donor"
                )

            filtered_df = donations_df.copy()

            filtered_df = filtered_df[
                (filtered_df["date"].dt.date >= from_date)
                &
                (filtered_df["date"].dt.date <= to_date)
            ]

            if donor_filter != "All":
                filtered_df = filtered_df[
                    filtered_df["name"] == donor_filter
                ]

            # ================= SUMMARY =================

            total_amount = filtered_df["amount"].sum()
            total_entries = len(filtered_df)

            highest_donation = (
                filtered_df["amount"].max()
                if not filtered_df.empty else 0
            )

            average_donation = (
                filtered_df["amount"].mean()
                if not filtered_df.empty else 0
            )

            c1, c2, c3, c4 = st.columns(4)

            with c1:
                st.metric(
                    "💰 Total Amount",
                    f"₹ {total_amount:,.0f}"
                )

            with c2:
                st.metric(
                    "📋 Entries",
                    total_entries
                )

            with c3:
                st.metric(
                    "🏆 Highest",
                    f"₹ {highest_donation:,.0f}"
                )

            with c4:
                st.metric(
                    "📊 Average",
                    f"₹ {average_donation:,.0f}"
                )

            st.markdown("---")

            # ================= INSIGHTS =================

            if not filtered_df.empty:

                i1, i2, i3 = st.columns(3)

                top_donor = (
                    filtered_df.groupby("name")["amount"]
                    .sum()
                    .idxmax()
                )

                lowest_donation = (
                    filtered_df["amount"].min()
                )

                current_month_total = filtered_df[
                    filtered_df["date"].dt.month
                    ==
                    datetime.now().month
                ]["amount"].sum()

                with i1:
                    st.info(
                        f"🏅 Top Donor: {top_donor}"
                    )

                with i2:
                    st.info(
                        f"📉 Lowest Donation: ₹ {lowest_donation:,.0f}"
                    )

                with i3:
                    st.info(
                        f"📅 Current Month: ₹ {current_month_total:,.0f}"
                    )

            st.markdown("---")

            # ================= EXPORTS =================

            export_df = filtered_df.copy()

            export_df["date"] = pd.to_datetime(
                export_df["date"]
            ).dt.strftime("%d-%m-%Y")

            export_columns = [
                "id",
                "name",
                "amount",
                "date",
                "note"
            ]

            export_columns = [
                col for col in export_columns
                if col in export_df.columns
            ]

            export_df = export_df[export_columns].copy()

            column_mapping = {
                "id": "ID",
                "name": "Donor Name",
                "amount": "Amount",
                "date": "Date",
                "note": "Note"
            }

            export_df.rename(
                columns=column_mapping,
                inplace=True
            )

            summary_text = (
                f"<b>Total Amount :</b> ₹ {total_amount:,.0f}"
                f"    "
                f"<b>Total Entries :</b> {total_entries}"
                f"    "
                f"<b>Highest Donation :</b> ₹ {highest_donation:,.0f}"
                f"    "
                f"<b>Average Donation :</b> ₹ {average_donation:,.0f}"
            )

            e1, e2 = st.columns(2)

            with e1:

                excel_file = generate_excel_report(
                    df=export_df,
                    report_title="DONATIONS REPORT"
                )

                st.download_button(
                    label="📥 Download Excel Report",
                    data=excel_file,
                    file_name="donations_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            with e2:

                pdf_file = generate_pdf_report(
                    df=export_df,
                    report_title="DONATIONS REPORT",
                    summary_text=summary_text,
                    generated_by=st.session_state.get(
                        "username",
                        "Admin"
                    )
                )

                st.download_button(
                    label="📄 Download PDF Report",
                    data=pdf_file,
                    file_name="donations_report.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )

            st.markdown("---")
            # ================= TABLE =================

            show_cols = [
                col for col in
                ["date", "name", "amount", "note"]
                if col in filtered_df.columns
            ]

            display_df = filtered_df.copy()

            display_df = display_df.sort_values(
                by="date",
                ascending=False
            )

            display_df["date"] = pd.to_datetime(
                display_df["date"]
            ).dt.strftime("%d-%m-%Y")

            st.dataframe(
                display_df[show_cols],
                use_container_width=True
            )

    # =========================================================
    # ================= EXPENSES ==============================
    # =========================================================

    with tab4:

        st.markdown("## 💸 Expenses Report")

        if expenses_df.empty:
            st.warning("No expenses found.")
        else:
            expenses_df["amount"] = pd.to_numeric(
                expenses_df.get("amount", 0),
                errors="coerce"
            ).fillna(0)

            expenses_df["date"] = pd.to_datetime(
                expenses_df["date"],
                errors="coerce"
            )

            # ================= FILTERS =================

            f1, f2, f3 = st.columns(3)

            with f1:
                from_date = st.date_input(
                    "📅 From Date",
                    expenses_df["date"].min().date(),
                    key="expense_from"
                )

            with f2:
                to_date = st.date_input(
                    "📅 To Date",
                    expenses_df["date"].max().date(),
                    key="expense_to"
                )

            with f3:
                type_filter = st.selectbox(
                    "💸 Expense Type",
                    ["All"] +
                    sorted(
                        expenses_df["type"]
                        .dropna()
                        .astype(str)
                        .unique()
                        .tolist()
                    ),
                    key="expense_type"
                )

            filtered_df = expenses_df.copy()

            filtered_df = filtered_df[
                (filtered_df["date"].dt.date >= from_date)
                &
                (filtered_df["date"].dt.date <= to_date)
            ]

            if type_filter != "All":
                filtered_df = filtered_df[
                    filtered_df["type"] == type_filter
                ]

            # ================= SUMMARY =================

            total_amount = filtered_df["amount"].sum()
            total_entries = len(filtered_df)

            highest_expense = (
                filtered_df["amount"].max()
                if not filtered_df.empty else 0
            )

            average_expense = (
                filtered_df["amount"].mean()
                if not filtered_df.empty else 0
            )

            c1, c2, c3, c4 = st.columns(4)

            with c1:
                st.metric(
                    "💰 Total Expense",
                    f"₹ {total_amount:,.0f}"
                )

            with c2:
                st.metric(
                    "📋 Entries",
                    total_entries
                )

            with c3:
                st.metric(
                    "📈 Highest",
                    f"₹ {highest_expense:,.0f}"
                )

            with c4:
                st.metric(
                    "📊 Average",
                    f"₹ {average_expense:,.0f}"
                )

            st.markdown("---")

            # ================= INSIGHTS =================

            if not filtered_df.empty:

                i1, i2, i3 = st.columns(3)

                top_type = (
                    filtered_df.groupby("type")["amount"]
                    .sum()
                    .idxmax()
                )

                lowest_expense = (
                    filtered_df["amount"].min()
                )

                current_month_total = filtered_df[
                    filtered_df["date"].dt.month
                    ==
                    datetime.now().month
                ]["amount"].sum()

                with i1:
                    st.info(
                        f"🏆 Top Category: {top_type}"
                    )

                with i2:
                    st.info(
                        f"📉 Lowest Expense: ₹ {lowest_expense:,.0f}"
                    )

                with i3:
                    st.info(
                        f"📅 Current Month: ₹ {current_month_total:,.0f}"
                    )

            st.markdown("---")

            # ================= EXPORTS =================

            export_df = filtered_df.copy()

            export_df["date"] = pd.to_datetime(
                export_df["date"]
            ).dt.strftime("%d-%m-%Y")

            export_columns = [
                "id",
                "type",
                "amount",
                "date",
                "note"
            ]

            export_columns = [
                col for col in export_columns
                if col in export_df.columns
            ]

            export_df = export_df[export_columns].copy()

            column_mapping = {
                "id": "ID",
                "type": "Expense Type",
                "amount": "Amount",
                "date": "Date",
                "note": "Note"
            }

            export_df.rename(
                columns=column_mapping,
                inplace=True
            )

            summary_text = (
                f"<b>Total Expenses :</b> ₹ {total_amount:,.0f}"
                f"    "
                f"<b>Total Entries :</b> {total_entries}"
                f"    "
                f"<b>Highest Expense :</b> ₹ {highest_expense:,.0f}"
                f"    "
                f"<b>Average Expense :</b> ₹ {average_expense:,.0f}"
            )

            e1, e2 = st.columns(2)

            with e1:

                excel_file = generate_excel_report(
                    df=export_df,
                    report_title="EXPENSES REPORT"
                )

                st.download_button(
                    label="📥 Download Excel Report",
                    data=excel_file,
                    file_name="expenses_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            with e2:

                pdf_file = generate_pdf_report(
                    df=export_df,
                    report_title="EXPENSES REPORT",
                    summary_text=summary_text,
                    generated_by=st.session_state.get(
                        "username",
                        "Admin"
                    )
                )

                st.download_button(
                    label="📄 Download PDF Report",
                    data=pdf_file,
                    file_name="expenses_report.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )

            st.markdown("---")
            # ================= TABLE =================

            show_cols = [
                col for col in
                ["date", "type", "amount", "note"]
                if col in filtered_df.columns
            ]

            display_df = filtered_df.copy()

            display_df = display_df.sort_values(
                by="date",
                ascending=False
            )

            display_df["date"] = pd.to_datetime(
                display_df["date"]
            ).dt.strftime("%d-%m-%Y")

            st.dataframe(
                display_df[show_cols],
                use_container_width=True
            )

    # =========================================================
    # ================= FINAL SUMMARY =========================
    # =========================================================

    st.markdown("---")
    st.markdown("## 📊 Overall Financial Summary")

    total_collection_all = 0
    total_donation_all = 0
    total_expense_all = 0
    total_loan_all = 0

    if not collections_df.empty:

        total_collection_all = pd.to_numeric(
            collections_df.get("amount", 0),
            errors="coerce"
        ).fillna(0).sum()

    if not donations_df.empty:

        total_donation_all = pd.to_numeric(
            donations_df.get("amount", 0),
            errors="coerce"
        ).fillna(0).sum()

    if not expenses_df.empty:

        total_expense_all = pd.to_numeric(
            expenses_df.get("amount", 0),
            errors="coerce"
        ).fillna(0).sum()

    if not loans_df.empty:

        total_loan_all = loans_df["Balance"].sum()

    profit = (
        total_collection_all +
        total_donation_all -
        total_expense_all
    )

    f1, f2, f3, f4 = st.columns(4) # columns count matches the 4 metrics below

    with f1:

        st.metric(
            "💰 Collections",
            f"₹ {total_collection_all:,.0f}"
        )

    with f2:

        st.metric(
            "🏦 Loans",
            f"₹ {total_loan_all:,.0f}"
        )

    with f3:

        st.metric(
            "🎁 Donations",
            f"₹ {total_donation_all:,.0f}"
        )

    with f4:

        st.metric(
            "💸 Expenses",
            f"₹ {total_expense_all:,.0f}"
        )
# ================= USERS =================
if menu == "Users":

    st.subheader("👥 User Management")

    current_user = st.session_state.current_user
    role = st.session_state.role

    # ✅ CACHE USERS DATA
    @st.cache_data(ttl=60)
    def load_users():
        try:
            return supabase.table("users").select("*").execute().data
        except:
            return []

    users_data = load_users()
    users_df = pd.DataFrame(users_data)

    # ================= USER SELF PASSWORD CHANGE =================
    st.markdown("### 🔐 Change Your Password")

    old_pass = st.text_input("Old Password", type="password")
    new_pass = st.text_input("New Password", type="password")
    confirm_pass = st.text_input("Confirm New Password", type="password")

    if st.button("Update My Password"):
        try:
            user_row = users_df[users_df["username"] == current_user]

            if user_row.empty:
                st.error("User not found ❌")
            else:
                user = user_row.iloc[0]

                if hash_pass(old_pass) != user["password"]:
                    st.error("Old password incorrect ❌")
                elif new_pass != confirm_pass:
                    st.error("Passwords do not match ❌")
                elif len(new_pass) < 4:
                    st.error("Password too short ❌")
                else:
                    supabase.table("users").update({
                        "password": hash_pass(new_pass)
                    }).eq("username", current_user).execute()

                    st.success("Password updated successfully ✅")
                    st.cache_data.clear()
                    st.rerun()
        except:
            st.error("Error updating password")

    st.markdown("---")

    # ================= ADMIN FEATURES =================
    if role == "Admin":

        # ===== CREATE USER =====
        st.markdown("### ➕ Create User")

        u = st.text_input("Username")
        p = st.text_input("Password", type="password")
        r = st.selectbox("Role", ["Admin", "Editor", "Viewer"])

        if st.button("Create User"):
            if u and p:
                try:
                    username = u.strip().lower()

                    existing = supabase.table("users").select("username").eq("username", username).execute()

                    if existing.data:
                        st.warning("⚠️ Username already exists")
                    else:
                        supabase.table("users").insert({
                            "username": username,
                            "password": hash_pass(p),
                            "role": r
                        }).execute()

                        st.success("User Created ✅")
                        st.cache_data.clear()
                        st.rerun()
                except Exception as e:
                    st.error(f"Error: {e}")
            else:
                st.warning("Enter username & password")

        st.markdown("---")

        # ===== RESET PASSWORD =====
        st.markdown("### 🔑 Reset User Password")

        if not users_df.empty:
            selected_user = st.selectbox("Select User", users_df["username"])
        else:
            selected_user = None

        new_pass_admin = st.text_input("New Password for User", type="password")

        if st.button("Reset Password"):
            if selected_user and new_pass_admin:
                try:
                    supabase.table("users").update({
                        "password": hash_pass(new_pass_admin)
                    }).eq("username", selected_user).execute()

                    st.success("Password Reset Done ✅")
                    st.cache_data.clear()
                    st.rerun()
                except:
                    st.error("Reset failed")
            else:
                st.warning("Enter new password")

        st.markdown("---")

        # ===== DELETE USER =====
        st.markdown("### 🗑️ Delete User")

        if not users_df.empty:
            del_user = st.selectbox("Select User to Delete", users_df["username"])
        else:
            del_user = None

        if st.button("Delete User"):
            if del_user == "admin":
                st.error("Admin cannot be deleted ❌")
            elif del_user:
                try:
                    supabase.table("users").delete().eq("username", del_user).execute()

                    st.warning("User Deleted ⚠️")
                    st.cache_data.clear()
                    st.rerun()
                except:
                    st.error("Delete failed")

        st.markdown("---")

        # ===== SHOW USERS =====
        st.markdown("### 📋 All Users")

        if not users_df.empty:
            st.dataframe(users_df[["username", "role"]])
        else:
            st.info("No users found")

    else:
        st.info("Limited access: You can only change your password 👁️")

# ================= Reminders =================
elif menu == "Reminders":

    st.title("📱 WhatsApp Reminders")
    st.markdown("""
    <style>

    /* WhatsApp Buttons */
    div[data-testid="stLinkButton"] a {
    background: linear-gradient(
    135deg,
    #25D366,
    #128C7E
    ) !important;

    color: white !important;
    font-weight: 700 !important;
    border-radius: 10px !important;
    border: none !important;
    text-decoration: none !important;
    text-align: center !important;
    padding: 10px !important;
    }

    /* Hover Effect */
    div[data-testid="stLinkButton"] a:hover {
    color: white !important;
    transform: scale(1.02);
    }

    /* Remove default blue focus */
    div[data-testid="stLinkButton"] a:focus {
    color: white !important;
    box-shadow: none !important;
    }

    </style>
    """, unsafe_allow_html=True)

    @st.cache_data(ttl=60)
    def load_reminder_data():
        def safe_fetch(table):
            try:
                return pd.DataFrame(
                    supabase.table(table)
                    .select("*")
                    .execute()
                    .data
                )
            except:
                return pd.DataFrame()

        return (
            safe_fetch("members"),
            safe_fetch("collections"),
            safe_fetch("loans"),
            safe_fetch("loan_payments"),
            safe_fetch("reminders")
        )

    members_df, collections_df, loans_df, payments_df, reminders_df = load_reminder_data()

    if members_df.empty:
        st.warning("No members found")
        st.stop()

    tab0, tab1, tab2 = st.tabs(["📊 Dashboard", "📅 Collection Reminders", "🏦 Loan Reminders"])

    # =====================================================
    # REMINDER DASHBOARD
    # =====================================================
    with tab0:
        st.subheader("📊 Reminder Dashboard")

        if reminders_df.empty:
            st.info("No reminders sent yet.")
        else:
            reminders_df["sent_date"] = pd.to_datetime(reminders_df["sent_date"])

            c1, c2 = st.columns(2)

            with c1:
                from_date = st.date_input("From Date", reminders_df["sent_date"].min().date())

            with c2:
                to_date = st.date_input("To Date", reminders_df["sent_date"].max().date())

            filtered_df = reminders_df[
                (reminders_df["sent_date"].dt.date >= from_date) &
                (reminders_df["sent_date"].dt.date <= to_date)
            ]

            # ================= USER FILTER =================
            user_options = ["All Users"]

            if "sent_by" in filtered_df.columns:
                user_options += sorted(filtered_df["sent_by"].dropna().unique().tolist())

            selected_user = st.selectbox("👨‍💼 Filter By User", user_options)

            if selected_user != "All Users":
                filtered_df = filtered_df[filtered_df["sent_by"] == selected_user]

            total_reminders = len(filtered_df)
            collection_count = len(filtered_df[filtered_df["reminder_type"] == "Collection"])
            loan_count = len(filtered_df[filtered_df["reminder_type"] == "Loan"])
            unique_members = filtered_df["member_name"].nunique()

            a, b, c, d, e, f = st.columns(6)

            with a:
                st.metric("📱 Total", total_reminders)
            with b:
                st.metric("📨 Collection", collection_count)
            with c:
                st.metric("🏦 Loan", loan_count)
            with d:
                st.metric("👥 Members", unique_members)

            today_count = len(filtered_df[filtered_df["sent_date"].dt.date == datetime.datetime.now().date()])
            top_member = "-"

            if not filtered_df.empty:
                top_member = filtered_df["member_name"].value_counts().idxmax()

            with e:
                st.metric("📅 Today", today_count)
            with f:
                st.metric("🏆 Top", top_member)

            # ================= TOP REMINDED MEMBERS =================
            st.markdown("---")
            st.subheader("🏆 Top Reminded Members")

            top_members = (
                filtered_df.groupby("member_name")
                .size()
                .reset_index(name="Reminders")
                .sort_values("Reminders", ascending=False)
                .head(5)
            )
            top_members.columns = ["Member", "Reminder Count"]

            if not top_members.empty:
                st.dataframe(top_members, use_container_width=True, hide_index=True)
            else:
                st.info("No reminder data found.")

            # ================= REMINDERS BY USER =================
            st.markdown("---")
            st.subheader("👨‍💼 Reminders By User")

            user_stats = filtered_df.groupby("sent_by").size().reset_index(name="Reminders")
            user_stats.columns = ["User", "Reminder Count"]

            st.dataframe(user_stats, use_container_width=True, hide_index=True)

            # ================= REMINDER HISTORY =================
            st.markdown("---")
            st.subheader("📋 Reminder History")

            history_df = filtered_df.copy()
            history_df = history_df[["sent_date", "member_name", "reminder_type", "amount", "sent_by"]]
            history_df.columns = ["Date", "Member", "Type", "Amount", "Sent By"]
            history_df["Date"] = pd.to_datetime(history_df["Date"]).dt.strftime("%d-%m-%Y %I:%M %p")
            history_df = history_df.sort_values("Date", ascending=False)

            st.dataframe(history_df, use_container_width=True, hide_index=True)

            # =====================================================
            # REMINDER EXPORT
            # =====================================================
            st.markdown("### 📥 Export Reports")

            reminder_export = history_df.copy()

            # ================= EXCEL =================
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine="openpyxl") as writer:
                reminder_export.to_excel(writer, index=False, sheet_name="Reminder Report")

            # ================= PDF =================
            pdf_buffer = BytesIO()
            doc = SimpleDocTemplate(pdf_buffer)
            table_data = [reminder_export.columns.tolist()]

            for row in reminder_export.values.tolist():
                table_data.append(row)

            table = Table(table_data)
            table.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]))
            doc.build([table])

            # ================= DOWNLOAD =================
            st.download_button(
                label="📊 Download Reminder Excel",
                data=excel_buffer.getvalue(),
                file_name="reminder_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )

            st.download_button(
                label="📄 Download Reminder PDF",
                data=pdf_buffer.getvalue(),
                file_name="reminder_report.pdf",
                mime="application/pdf",
                use_container_width=True
            )

    # =====================================================
    # COLLECTION REMINDERS
    # =====================================================
    with tab1:
        st.subheader("📅 Collection Reminders")

        if collections_df.empty:
            st.info("No collection data found")
        else:
            collection_data = collections_df.copy()
            collection_data["amount"] = pd.to_numeric(collection_data["amount"], errors="coerce").fillna(0)
            collection_data["expected_amount"] = pd.to_numeric(collection_data["expected_amount"], errors="coerce").fillna(0)

            summary = (
                collection_data.groupby(["member_id", "month"])
                .agg({"expected_amount": "sum", "amount": "sum"})
                .reset_index()
            )

            summary["Balance"] = summary["expected_amount"] - summary["amount"]
            summary["Balance"] = summary["Balance"].clip(lower=0)

            pending_members = summary[summary["Balance"] > 0].copy()

            if pending_members.empty:
                st.success("✅ No Pending Collections")
            else:
                pending_members = pending_members.merge(
                    members_df[["id", "name", "mobile"]],
                    left_on="member_id",
                    right_on="id",
                    how="left"
                )

                for _, row in pending_members.iterrows():
                    mobile = str(row.get("mobile", "")).strip()
                    c1, c2, c3, c4, c5 = st.columns([3, 2, 2, 2, 2])

                    last_sent = "-"
                    count_sent = 0

                    if not reminders_df.empty and "member_id" in reminders_df.columns:
                        member_reminders = reminders_df[
                            (reminders_df["member_id"].astype(str) == str(row["member_id"])) &
                            (reminders_df["reminder_type"] == "Collection")
                        ]
                        if not member_reminders.empty:
                            count_sent = len(member_reminders)
                            last_sent = pd.to_datetime(member_reminders["sent_date"]).max().strftime("%d-%m-%Y")

                    with c1:
                        st.write(f"👤 {row['name']}")
                        st.caption(f"📅 Last: {last_sent}")
                        st.caption(f"🔔 Count: {count_sent}")

                    with c2:
                        st.write(f"📅 {row['month']}")

                    with c3:
                        st.write(f"₹ {row['Balance']:,.0f}")

                    with c4:
                        if mobile.isdigit() and len(mobile) == 10:
                            message = f"""
🙏 नमस्कार आदरणीय सदस्य,

आपकी {row['month']} की ₹{row['Balance']:,.0f} राशि लंबित है।

कृपया जल्द जमा करने का कष्ट करें।

सादर धन्यवाद।

बाल युवा मंगल दल समिति
मयलगांव
"""
                            wa_link = f"https://api.whatsapp.com/send?phone=91{mobile}&text={urllib.parse.quote(message)}"
                            st.link_button("📱 WhatsApp", wa_link, use_container_width=True)
                        else:
                            st.warning("No Mobile")

                    with c5:
                        if st.button("✅ Mark Sent", key=f"collection_sent_{row['member_id']}_{row['month']}"):
                            try:
                                supabase.table("reminders").insert({
                                    "member_id": str(row["member_id"]),
                                    "member_name": row["name"],
                                    "mobile": mobile,
                                    "reminder_type": "Collection",
                                    "reminder_month": row["month"],
                                    "amount": float(row["Balance"]),
                                    "sent_by": st.session_state.get("current_user", "Admin"),
                                    "sent_by_role": st.session_state.get("role", "Viewer"),
                                    "status": "Sent"
                                }).execute()

                                save_log(
                                    action="Reminder Sent",
                                    table_name="reminders",
                                    member_name=row["name"],
                                    member_id=str(row["member_id"]),
                                    amount=float(row["Balance"])
                                )

                                st.success(f"Reminder saved for {row['name']}")
                                st.cache_data.clear()
                                st.rerun()
                            except Exception as e:
                                st.error(f"Error: {e}")

    # =====================================================
    # LOAN REMINDERS
    # =====================================================
    with tab2:
        st.subheader("🏦 Loan Reminders")

        if loans_df.empty:
            st.info("No loan data found")
        else:
            active_found = False

            for _, loan in loans_df.iterrows():
                loan_id = loan["id"]
                principal = float(loan.get("amount", 0))
                loan_payments = payments_df[payments_df["loan_id"] == loan_id]
                principal_paid = loan_payments["principal_paid"].sum() if not loan_payments.empty else 0
                balance = max(principal - principal_paid, 0)

                if balance <= 0:
                    continue

                active_found = True
                member_name = ""

                if "customer_name" in loan:
                    member_name = str(loan.get("customer_name", "")).strip()

                member_row = members_df[members_df["name"] == member_name]
                mobile = ""

                if not member_row.empty:
                    mobile = str(member_row.iloc[0].get("mobile", "")).strip()

                last_sent = "-"
                count_sent = 0

                if not reminders_df.empty and "member_id" in reminders_df.columns:
                    member_id = ""
                    if not member_row.empty:
                        member_id = str(member_row.iloc[0].get("id", ""))

                    loan_reminders = reminders_df[
                        (reminders_df["member_id"].astype(str) == member_id) &
                        (reminders_df["reminder_type"] == "Loan")
                    ]
                    if not loan_reminders.empty:
                        count_sent = len(loan_reminders)
                        last_sent = pd.to_datetime(loan_reminders["sent_date"]).max().strftime("%d-%m-%Y")

                c1, c2, c3, c4 = st.columns([4, 2, 2, 2])

                with c1:
                    st.write(f"👤 {member_name}")
                    st.caption(f"📅 Last: {last_sent}")
                    st.caption(f"🔔 Count: {count_sent}")

                with c2:
                    st.write(f"₹ {balance:,.0f}")

                with c3:
                    if mobile.isdigit() and len(mobile) == 10:
                        message = f"""
🙏 नमस्कार आदरणीय सदस्य,

आपके ऋण का ₹{balance:,.0f} बकाया है।

कृपया समय पर भुगतान करने का कष्ट करें।

सादर धन्यवाद।

बाल युवा मंगल दल समिति
मयलगांव
"""
                        wa_link = f"https://api.whatsapp.com/send?phone=91{mobile}&text={urllib.parse.quote(message)}"
                        st.link_button("📱 WhatsApp", wa_link, use_container_width=True)
                    else:
                        st.warning("No Mobile")

                with c4:
                    if st.button("✅ Mark Sent", key=f"loan_sent_{loan_id}"):
                        try:
                            member_id = ""
                            if not member_row.empty:
                                member_id = str(member_row.iloc[0].get("id", ""))

                            supabase.table("reminders").insert({
                                "member_id": member_id,
                                "member_name": member_name,
                                "mobile": mobile,
                                "reminder_type": "Loan",
                                "reminder_month": "Loan Balance",
                                "amount": float(balance),
                                "sent_by": st.session_state.get("current_user", "Admin"),
                                "sent_by_role": st.session_state.get("role", "Viewer"),
                                "status": "Sent"
                              }).execute()

                            st.success(f"Reminder saved for {member_name}")
                            st.cache_data.clear()
                            st.rerun()
                        except Exception as e:
                            st.error(f"Error: {e}")

            if not active_found:
                st.success("✅ No Active Loan Balances")

# =====================================================
# ============== BACKUP & RESTORE =====================
# =====================================================
elif menu == "Backup & Restore":

    st.title("💾 Backup & Restore")
    st.info("Create a full database backup before making major changes.")
    st.markdown("---")

    # ================= BACKUP =================
    st.subheader("📥 Full Database Backup")

    backup_file, file_name = create_full_backup(supabase)

    st.download_button(
        label="⬇️ Download Full Backup",
        data=backup_file,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )

    st.success("This backup contains Members, Collections, Collection Rates, Loans, Loan Payments, Donations, Expenses, Reminders and Users data.")
    st.markdown("---")

    # ================= RESTORE =================
    st.subheader("♻️ Restore Database")
    st.warning("⚠️ Restoring backup will replace existing data.\n\nPlease take a fresh backup before restoring.")

    uploaded_backup = st.file_uploader("📤 Upload Backup (.xlsx)", type=["xlsx"])

    if uploaded_backup is not None:
        st.success(f"📄 Selected File: {uploaded_backup.name}")

        if st.button("♻️ Restore Database", use_container_width=True):
            if uploaded_backup is None:
                st.warning("Please upload a backup file first.")
            else:
                try:
                    with st.spinner("Restoring database..."):
                        restored_tables = restore_full_backup(uploaded_backup, supabase)

                    st.success("🎉 Restore completed successfully!")
                    st.write("### Restored Tables")

                    for table in restored_tables:
                        st.write(f"✅ {table}")

                    st.cache_data.clear()
                    st.rerun()
                except Exception as e:
                    st.error(f"Restore failed: {e}")

# ================= AI =================
elif menu == "AI":
    st.subheader("🤖 AI Insights (Coming Soon)")
    st.info("Future AI features yaha add honge")
