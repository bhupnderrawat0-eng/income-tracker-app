from io import BytesIO
from datetime import datetime

import pandas as pd

# ================= EXCEL IMPORTS =================

from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, PatternFill

# ================= PDF IMPORTS =================

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image
)

from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_CENTER


# =====================================================
# =============== COMMON EXCEL REPORT =================
# =====================================================

def generate_excel_report(
    df,
    report_title="REPORT"
):

    excel_buffer = BytesIO()

    with pd.ExcelWriter(
        excel_buffer,
        engine="openpyxl"
    ) as writer:

        # ================= DATA =================

        df.to_excel(
            writer,
            index=False,
            sheet_name="Report",
            startrow=7
        )

        worksheet = writer.sheets["Report"]

        # ================= LOGO =================

        try:
            logo = ExcelImage("logo.png")
            logo.width = 90
            logo.height = 90
            worksheet.add_image(logo, "A1")
        except:
            pass

        # ================= HEADER =================

        worksheet.merge_cells("C1:H1")
        worksheet["C1"] = "बाल युवा मंगलदल समिति"
        worksheet["C1"].font = Font(
            size=18,
            bold=True,
            color="F8D568"
        )
        worksheet["C1"].alignment = Alignment(
            horizontal="center"
        )

        worksheet.merge_cells("C2:H2")
        worksheet["C2"] = "मयलगांव"
        worksheet["C2"].font = Font(
            size=14,
            bold=True,
            color="EFD58A"
        )
        worksheet["C2"].alignment = Alignment(
            horizontal="center"
        )

        worksheet.merge_cells("C3:H3")
        worksheet["C3"] = (
            "हमारा गांव • हमारी पहचान • हमारा अभियान"
        )
        worksheet["C3"].font = Font(
            size=12,
            italic=True
        )
        worksheet["C3"].alignment = Alignment(
            horizontal="center"
        )

        worksheet.merge_cells("A5:H5")
        worksheet["A5"] = report_title
        worksheet["A5"].font = Font(
            size=16,
            bold=True
        )
        worksheet["A5"].alignment = Alignment(
            horizontal="center"
        )

        worksheet.merge_cells("A6:H6")
        worksheet["A6"] = (
            f"Generated On : "
            f"{datetime.now().strftime('%d-%m-%Y %I:%M %p')}"
        )
        worksheet["A6"].alignment = Alignment(
            horizontal="center"
        )

        # ================= TABLE HEADER STYLE =================

        for cell in worksheet[8]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="D9D9D9"
            )
            cell.alignment = Alignment(horizontal="center")

        # ================= FIXED COLUMN WIDTH =================

        worksheet.column_dimensions["A"].width = 18
        worksheet.column_dimensions["B"].width = 25
        worksheet.column_dimensions["C"].width = 18
        worksheet.column_dimensions["D"].width = 20
        worksheet.column_dimensions["E"].width = 18
        worksheet.column_dimensions["F"].width = 22
        worksheet.column_dimensions["G"].width = 15
        worksheet.column_dimensions["H"].width = 15
        worksheet.column_dimensions["I"].width = 15
        worksheet.column_dimensions["J"].width = 15

    excel_buffer.seek(0)

    return excel_buffer


# =====================================================
# ================= COMMON PDF REPORT =================
# =====================================================

def generate_pdf_report(
    df,
    report_title="REPORT",
    summary_text="",
    generated_by="Admin"
):

    pdf_buffer = BytesIO()

    doc = SimpleDocTemplate(
        pdf_buffer,
        topMargin=20,
        bottomMargin=20
    )

    styles = getSampleStyleSheet()

    title_style = styles["Heading1"]
    title_style.alignment = TA_CENTER

    center_style = styles["BodyText"]
    center_style.alignment = TA_CENTER

    elements = []

    # ================= LOGO =================

    try:
        logo = Image("logo.png")
        logo.drawHeight = 140
        logo.drawWidth = 140
        logo.hAlign = "CENTER"
        elements.append(logo)
    except:
        pass

    # ================= HEADER =================

    elements.append(
        Paragraph(
            "<b>Bal Yuva Mangal Dal Samiti</b>",
            title_style
        )
    )

    elements.append(
        Paragraph(
            "<b>Mayalgaon</b>",
            center_style
        )
    )

    elements.append(
        Paragraph(
            "Hamara Gaon • Hamari Pehchan • Hamara Abhiyan",
            center_style
        )
    )

    elements.append(Spacer(1, 10))

    elements.append(
        Paragraph(
            f"<b>{report_title}</b>",
            title_style
        )
    )

    elements.append(
        Paragraph(
            f"Generated On : {datetime.now().strftime('%d-%m-%Y %I:%M %p')}",
            center_style
        )
    )

    elements.append(Spacer(1, 10))

    # ================= SUMMARY =================

    if summary_text:
        elements.append(
            Paragraph(
                summary_text,
                center_style
            )
        )

        elements.append(Spacer(1, 15))

    # ================= CLEAN STATUS =================

    clean_df = df.copy()

    if "Status" in clean_df.columns:
        clean_df["Status"] = (
            clean_df["Status"]
            .astype(str)
            .str.replace("⚠️", "", regex=False)
            .str.replace("✅", "", regex=False)
            .str.strip()
        )

    # ================= TABLE =================

    table_data = [clean_df.columns.tolist()]
    table_data += clean_df.values.tolist()

    table = Table(table_data)

    table.setStyle(
        TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0),
             colors.HexColor("#EFD58A")),

            ("TEXTCOLOR", (0, 0), (-1, 0),
             colors.black),

            ("GRID", (0, 0), (-1, -1),
             1, colors.black),

            ("FONTNAME", (0, 0), (-1, 0),
             "Helvetica-Bold"),

            ("FONTSIZE", (0, 0), (-1, -1),
             8),

            ("ALIGN", (0, 0), (-1, -1),
             "CENTER"),
        ])
    )

    elements.append(table)

    elements.append(Spacer(1, 25))

    # ================= FOOTER =================

    elements.append(
        Paragraph(
            "<b>Bal Yuva Mangal Dal Samiti</b>",
            center_style
        )
    )

    elements.append(
        Paragraph(
            "Mayalgaon",
            center_style
        )
    )

    elements.append(
        Paragraph(
            f"Report Generated By : {generated_by}",
            center_style
        )
    )

    elements.append(
        Paragraph(
            "* End of Report *",
            center_style
        )
    )

    doc.build(elements)

    pdf_buffer.seek(0)

    return pdf_buffer
